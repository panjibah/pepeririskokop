from tkinter.messagebox import NO
from django.db.models.query import prefetch_related_objects
from django.db.models.functions import Length
from django.http.response import JsonResponse
from django.shortcuts import render
from django.shortcuts import redirect
from django.db import connection
from ledMaster.models import LedHistoris
from .models import Cabang, Coa, Kasus, Led,Laporan,LedLaporan,KategoriKejadian,Penyebab, StateChange, SummaryKejadian, Unit,BusinessLine,Produk,KategoriLaporan,Perubahan,Journal
from django.db.models import Q
import json
import os
from datetime import datetime
from django.http import HttpResponse
from django.core import serializers
from django.core.serializers.json import DjangoJSONEncoder
from django.http import JsonResponse
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment 
import pandas as pd
import pyexcel as px
from django.core.files import File
from django.contrib.staticfiles import finders
from functools import reduce
import operator
# from pytz import timezone
from django.utils import timezone
from .filters import SearchFilter
from django.core.paginator import Paginator,EmptyPage
from deepdiff import DeepDiff,extract
from pprint import pprint
# Create your views here.


class LazyEncoder(DjangoJSONEncoder):
    def default(self, obj):
        if isinstance(obj, YourCustomType):
            return str(obj)
        return super().default(obj)
    
def listLed(request):
    data = Led.objects.all()
    
    context= {
        'data'  : data
    }
    
    return render(request,'led/listLed.html',context)

def dummy(request):
    data = Led.objects.all()
    for i in data:
        i.delete()

def createReport(request):
    if request.POST:
        #get form data
        judul = request.POST.get('judul','')
        periode_awal = request.POST.get('start_prd','')
        periode_akhir = request.POST.get('end_prd','')
        kategoriLp = request.POST.get('kategori','')
        katlp = KategoriLaporan.objects.get(nama = kategoriLp)
        print(type(katlp))
        
        #validate form data
        
        #insert form data
        newLaporan = Laporan(
            nama = judul,
            periode_awal = periode_awal,
            periode_akhir = periode_akhir,
            kategori= katlp,
            creator = "Panji",      
        )
        newLaporan.save()
        
        print(newLaporan.periode_awal)
        print(newLaporan.periode_akhir)
        getNewLaporan = Laporan.objects.latest('id')
        print(type(getNewLaporan))
        #get led from ledlaporan
        rawquery1 = ("SELECT distinct l.id,l.nomor_kasus"
                    +" FROM led_ledlaporan lp join led_led l ON l.id = lp.led_laporan_id" 
                    +" WHERE l.tanggal_input between to_timestamp " + "('"+ periode_awal +"'" +",'YYYY-MM-DD')"+ "and to_timestamp" + "('"+ periode_akhir +"'" +",'YYYY-MM-DD')")
        
        # rawquery1 = ("SELECT distinct l.id,l.nomor_kasus"
        #             +" FROM led_kasus_laporan lk join led_led l ON l.id = lp.led_laporan_id" 
        #             +" WHERE l.tanggal_input between to_timestamp " + "('"+ periode_awal +"'" +",'YYYY-MM-DD')"+ "and to_timestamp" + "('"+ periode_akhir +"'" +",'YYYY-MM-DD')")
        
        print(rawquery1)
        usedLed = Led.objects.raw(rawquery1)
        for data in usedLed:
            print("s")
            insertLed=LedLaporan(
                laporan          = getNewLaporan,
                led_laporan = Led.objects.get(id = data.id),
            )
            insertLed.save()
        
        #get ked exclude from ledLaporan/new led that never processed
        # rawquery2 = ("SELECT distinct l.id,l.nomor_kasus"
        #             +" FROM led_led l"
        #             +" WHERE l.tanggal_input between to_timestamp " 
        #             +"('"+ periode_awal +"'" +",'YYYY-MM-DD')"+ "and to_timestamp" + "('"+ periode_akhir +"'" +",'YYYY-MM-DD')"
        #             +" and l.tanggal_update is NULL and"  
        #             +" NOT EXISTS (SELECT l2.led_laporan_id"
        #                             +" from led_ledlaporan l2 WHERE  l2.led_laporan_id = l.id )")
        #get ked exclude from ledLaporan/new led that never processed
        rawquery2 = ("SELECT  l.id "
                    +" FROM led_led l "
                    +" left join led_ledlaporan lp "
                    +" ON l.id = lp.led_laporan_id "
                    +" WHERE l.tanggal_input between to_timestamp ('2021-01-08','YYYY-MM-DD')and to_timestamp('2022-04-30','YYYY-MM-DD') and lp.id ISNULL ")
        
        print(rawquery2)
        notUsedLed = Led.objects.raw(rawquery2)
        for data in notUsedLed:
            
            insertLed=LedLaporan(
                laporan          = getNewLaporan,
                led_laporan = Led.objects.get(id = data.id),
            )
            insertLed.save()
        # getLed=Led.objects.filter(tanggal_input__range =(periode_awal,periode_akhir))
        # print(getLed.count())
        # for data in getLed:
        #     print("s")
        #     insertLed=LedLaporan(
        #         laporan          = getNewLaporan,
        #         led_laporan = Led.objects.get(id = data.id),
        #     )
        #     insertLed.save()
        kategori = KategoriLaporan.objects.all()
        context = {
            'kategori':kategori,
        }
        
        
        
        # return redirect('listLedReport',getNewLaporan.id)
        return render(request,'led/createRpt/createRpt.html',context)
        
    else:
        kategori = KategoriLaporan.objects.all()
        print(kategori.count())
        context = {
            "kategori":kategori,
        }
        
        return render(request,'led/createRpt/createRpt.html',context)
    
def listReport(request):
    laporan = Laporan.objects.all().order_by('created_at')

    # laporanBaru = Laporan.objects.latest('id')
    
    # logLaporan=json.dumps(laporanBaru,default=str)
    
    # qs_json = serializers.serialize('json', laporan)
    
    # # print(type(logLaporan))

    # # print(qs_json)
    
    # return JsonResponse(json.loads(qs_json),safe=False)
    context ={
        "laporan" :laporan
    }
    return render(request,'led/createRpt/listLaporan.html',context)

def testJson(request):
    laporanBaru = Laporan.objects.latest('id')
    led = Led.objects.all()
    kategori = KategoriKejadian.objects.all()
    penyebab =Penyebab.objects.all()
    unit = Unit.objects.all()
    cabang = Cabang.objects.all()
    bisnis =  BusinessLine.objects.all()
    produk =Produk.objects.all()
    

    
    logLaporan=json.dumps(laporanBaru,default=str)
    qs_json = serializers.serialize('json', [led],)
    print(type(qs_json))
    # print(type(logLaporan))
    # print(qs_json)
    
    return JsonResponse(json.loads(qs_json),safe=False)
    
    
def listLedReport(request,rptId):
    getLaporan = Laporan.objects.get(id = rptId)
    data = LedLaporan.objects.select_related('laporan','led_laporan','led_laporan__coa_biaya','led_laporan__kode_produk','led_laporan__kode_ktg_kejadian','led_laporan__kode_penyebab','led_laporan__business_line').filter(laporan = getLaporan)
    #bikin lemot manggil chekc tiap query
    kerugian = getKerugian(request,rptId)
    sumRevLed(request,getLaporan)
    # chkKategorixTipe(request,getLaporan)
    checkSmr(request,data)
    checkStatus(request,data)
    print(rptId)
    #get new data 
    #check kategori kejadian dan tipe led
    kknled = LedLaporan.objects.select_related('led_laporan','laporan').filter(Q(led_laporan__chk_kknled = True)&Q(laporan = getLaporan)).count()
    print(kknled)
    #summary kosong
    summNull = LedLaporan.objects.select_related('led_laporan','laporan').filter(Q(led_laporan__summary_kejadian = None)&Q(laporan = getLaporan)).count()
    # check sumary tidak terdapa
    summCount = LedLaporan.objects.select_related('led_laporan','laporan').filter(Q(led_laporan__chk_summ = True)&Q(laporan = getLaporan)).count()
    # getLed=Led.objects.filter(tanggal_kejadian__range =(laporan_id__periode_awal,laporan_id__periode_akhir))
    
    #perlu di review
    summCount = LedLaporan.objects.select_related('led_laporan','laporan').filter(Q(led_laporan__chk_summ = True)&Q(laporan = getLaporan)).count()
    
    # getNewLed = Led.objects.filter(tanggal_kejadian__range =(getLaporan.periode_awal,getLaporan.periode_akhir)).exclude(nomor_kasus__in = LedLaporan.objects.filter(Q(laporan_id = getLaporan)).values('nomor_kasus'))
    getNewLed = Led.objects.filter(tanggal_input__range =(getLaporan.periode_awal,getLaporan.periode_akhir)).exclude(nomor_kasus__in = LedLaporan.objects.filter(Q(laporan_id = getLaporan.id)).values('led_laporan_id__nomor_kasus'))
    perluDihapus= LedLaporan.objects.select_related('laporan','led_laporan').filter(laporan = getLaporan).filter(led_laporan__hasil_review__icontains = "Perlu Dihapus").count()
    tipeKerugian = ['Total Kerugian Aktual Tertinggi','Total Kerugian Potensial Tertinggi','Kerugian Aktual Tertinggi','Kerugian Potensial Tertinggi']

    zipList = zip(kerugian,tipeKerugian)

    print(getNewLed.count())
    rawNewData = ('Select *'
                    +  'from led_led l'
                    +   'where (l.tanggal_kejadian BETWEEN "2003-01-18" AND "2003-01-18") and (l.nomor_kasus in '
					+												  '(SELECT nomor_kasus '
					+												  	'FROM led_ledlaporan ll'
					+												  	'WHERE ll.tanggal_kejadian BETWEEN "2003-01-18" AND "2003-01-18"))')


    #get simmiliar/dupliacate data
    # getDuplicateData =  Led.objects.filter(Q(tanggal_kejadian__range =(getLaporan.periode_awal,getLaporan.periode_akhir))).filter(nomor_kasus__in = LedLaporan.objects.filter(Q(laporan_id = getLaporan)).values('nomor_kasus')).distinct('nama_pembuat')
    rawDupData = (  'Select *'
                    +   'from led_led l'
                    +   'where (l.tanggal_kejadian BETWEEN "2003-01-18" AND "2003-01-18") and (l.nomor_kasus in '
					+												  '(SELECT nomor_kasus '
					+												  	'FROM led_ledlaporan ll'
					+												  	'WHERE ll.tanggal_kejadian BETWEEN "2003-01-18" AND "2003-01-18"))')

    
    #perubahan data 
    
    perubahanData = Perubahan.objects.raw("select per.* from led_ledlaporan ll join led_led l on ll.led_laporan_id = l.id"
                                          +" join led_perubahan  per ON per.nomor_kasus_id = l.id "
                                        +" where exists ( select ll2.led_laporan_id"
			                            +" from led_ledlaporan ll2 where ll2.led_laporan_id = per.nomor_kasus_id and ll2.laporan_id ="+str(getLaporan.id)+" )")
    
    context ={
        "data" :data,
        "sizeData":len(data),
        "laporan":getLaporan,
        "getNewLed":getNewLed,
        "kknCount":kknled,
        "summNull":summNull,
        "summCount":summCount,
        "perluDihapus":perluDihapus,
        "kerugian":zipList,
    
    }

    return render(request,'led/ledRpt/ledReport.html',context)

def kerugian(request,rptId,tipe):
    
    #get kategori
    
    ktg = request.GET.get("kerugian",'')
    getLaporan = Laporan.objects.get(id = rptId)
    #if condition
    if(tipe == 1 or tipe == 2):
        #query total kerugian list led
        data = LedLaporan.objects.select_related('laporan','led_laporan','led_laporan__coa_biaya','led_laporan__kode_produk','led_laporan__kode_ktg_kejadian','led_laporan__kode_penyebab','led_laporan__business_line').filter(laporan = getLaporan).filter(led_laporan__kode_ktg_kejadian__nama_kategori3__icontains = ktg )
    else:
        data = LedLaporan.objects.select_related('laporan','led_laporan','led_laporan__coa_biaya','led_laporan__kode_produk','led_laporan__kode_ktg_kejadian','led_laporan__kode_penyebab','led_laporan__business_line').filter(laporan = getLaporan).order_by('-led_laporan__kerugian_aktual')[:1]
    

    
    context ={
        "data" :data,
        "laporan":getLaporan,
       
        # "getDuplicateData":getDuplicateData,
    }
    #return
    return render(request,'led/summRpt/totalKerugian.html',context)
    
    
def summaryReport(request,rptId):
    getLaporan = Laporan.objects.get(id = rptId)
    data = ""
    #check id query paramas
    #query
    #if totalData
    summCat= request.GET.get("summ",'')
    print(summCat)
    if(summCat == "totalData"):
        data = LedLaporan.objects.select_related('laporan','led_laporan','led_laporan__coa_biaya','led_laporan__kode_produk','led_laporan__kode_ktg_kejadian','led_laporan__kode_penyebab','led_laporan__business_line').filter(laporan = getLaporan)
    elif(summCat == "belumReview"):
        print("masuk 2")
        data = LedLaporan.objects.select_related('laporan','led_laporan','led_laporan__coa_biaya','led_laporan__kode_produk','led_laporan__kode_ktg_kejadian','led_laporan__kode_penyebab','led_laporan__business_line').filter(laporan = getLaporan).filter(led_laporan__review_user__icontains = "Belum Review")
    elif(summCat == "save"):
      
        data = LedLaporan.objects.select_related('laporan','led_laporan','led_laporan__coa_biaya','led_laporan__kode_produk','led_laporan__kode_ktg_kejadian','led_laporan__kode_penyebab','led_laporan__business_line').filter(laporan = getLaporan).filter(led_laporan__review_user__icontains = "save")
    elif(summCat == "submit"):
      
        data = LedLaporan.objects.select_related('laporan','led_laporan','led_laporan__coa_biaya','led_laporan__kode_produk','led_laporan__kode_ktg_kejadian','led_laporan__kode_penyebab','led_laporan__business_line').filter(laporan = getLaporan).filter(led_laporan__review_user__icontains = "submit")
    elif(summCat == "kkTdkSesuai"):
      
        data = LedLaporan.objects.select_related('laporan','led_laporan','led_laporan__coa_biaya','led_laporan__kode_produk','led_laporan__kode_ktg_kejadian','led_laporan__kode_penyebab','led_laporan__business_line').filter(laporan = getLaporan).filter(led_laporan__chk_kknled = True)
    elif(summCat == "summKosong"):
      
        data = LedLaporan.objects.select_related('laporan','led_laporan','led_laporan__coa_biaya','led_laporan__kode_produk','led_laporan__kode_ktg_kejadian','led_laporan__kode_penyebab','led_laporan__business_line').filter(laporan = getLaporan).filter(led_laporan__chk_summ = True)
    elif(summCat == "revKosong"):
      
        data = LedLaporan.objects.select_related('laporan','led_laporan','led_laporan__coa_biaya','led_laporan__kode_produk','led_laporan__kode_ktg_kejadian','led_laporan__kode_penyebab','led_laporan__business_line').filter(laporan = getLaporan).filter(led_laporan__hasil_review__lt = 1)
    elif(summCat == "dihapus"):
      
        data = LedLaporan.objects.select_related('laporan','led_laporan','led_laporan__coa_biaya','led_laporan__kode_produk','led_laporan__kode_ktg_kejadian','led_laporan__kode_penyebab','led_laporan__business_line').filter(laporan = getLaporan).filter(led_laporan__hasil_review__icontains = "Perlu Dihapus")
    else:
        
        print("gak masu")
        data=""
    #if 2
    
    context ={
        "data" :data,
        "laporan":getLaporan,
       
        # "getDuplicateData":getDuplicateData,
    }
    #return
    return render(request,'led/summRpt/totalData.html',context)
#menghitung total kerugian dan kerugian pada setiap katagori
def getKerugian(request,parRptId):
    rptId= str(parRptId)
    krg = Laporan.objects.raw("(SELECT  l.id,kk.nama_kategori1,sum(ld.kerugian_aktual)"
                        +" FROM led_led ld left join led_ledlaporan ll on ld.id = ll.led_laporan_id left join led_laporan l on l.id = ll.laporan_id left join led_kategorikejadian kk on kk.id = ld.kode_ktg_kejadian_id"
                        +" where l.id = " +rptId 
                        +" group by kk.nama_kategori1,l.id"
                        +" order by sum(ld.kerugian_aktual) desc"
                        +" limit 1)"
                        +" union all"
                        +" (SELECT  l.id,kk.nama_kategori1,sum(ld.kerugian_potensial)"
                        +" FROM led_led ld left join led_ledlaporan ll on ld.id = ll.led_laporan_id left join led_laporan l on l.id = ll.laporan_id left join led_kategorikejadian kk on kk.id = ld.kode_ktg_kejadian_id"
                        +" where l.id = "+rptId
                        +" group by kk.nama_kategori1,l.id"
                        +" order by sum(ld.kerugian_potensial) desc"
                        +" limit 1)"
                        +" union all"
                        +" (SELECT  l.id,kk.nama_kategori1,ld.kerugian_aktual"
                        +" FROM led_led ld left join led_ledlaporan ll on ld.id = ll.led_laporan_id left join led_laporan l on l.id = ll.laporan_id left join led_kategorikejadian kk on kk.id = ld.kode_ktg_kejadian_id"
                        +" where ld.kerugian_aktual IS NOT NULL and  l.id = "+rptId
                        +" order by(ld.kerugian_aktual) desc "
                        +" limit 1)"
                        +" union all"
                        +" (SELECT  l.id,kk.nama_kategori1,ld.kerugian_potensial"
                        +" FROM led_led ld left join led_ledlaporan ll on ld.id = ll.led_laporan_id left join led_laporan l on l.id = ll.laporan_id left join led_kategorikejadian kk on kk.id = ld.kode_ktg_kejadian_id"
                        +" where l.id = "+rptId
                        +" order by (ld.kerugian_potensial) desc"
                        +" limit 1)" )
    return krg
    
#menampilkan perubahan data yang terjadi pada laporan
def showPerubahanLaporan(request,lapId):
    getLaporan = Laporan.objects.get(id = lapId)
    perubahanData = Perubahan.objects.raw("select distinct per.*,l.nomor_kasus as no_kasus from led_ledlaporan ll join led_led l on ll.led_laporan_id = l.id"
                                          +" join led_perubahan  per ON per.nomor_kasus_id = l.id "
                                        +" where exists ( select ll2.led_laporan_id"
			                            +" from led_ledlaporan ll2 where ll2.led_laporan_id = per.nomor_kasus_id and ll2.laporan_id ="+str(getLaporan.id)+" ) order by per.created_at desc")
    
    context={
        'perubahanData' : perubahanData,
        'laporan':getLaporan, 
        
    }
    return render(request,'led/perubahan/perubahanData.html',context)

#memasukkan perubahan data yang didaptkan saat menjalankan function untuk membandingkan data baru dan dat yang ad apada laporan
def create_state(request,stateDict):
    print("kwargs")
    result = ""
    # Iterating over the keys of the Python kwargs dictionary
    print(stateDict['nomor_kasus'])
    print(stateDict['nomor_kasus2'])
    print(stateDict['nama_kolom'])
    print(stateDict['value'])
    
    stateChange = StateChange(
        nomor_kasus = Led.objects.get(id = stateDict['nomor_kasus']),
        nomor_kasus2 = Led.objects.get(id = stateDict['nomor_kasus2']),
        nama_kolom =stateDict['nama_kolom'],
        value = stateDict['value'],
        status = 1
    )
    stateChange.save()

#menampilkan table statechange
def getStateChange(request):
    ledUpdate = Led.objects.exclude(tanggal_update = None)
    ledUpdate2 = LedLaporan.objects.values('led')
    print(ledUpdate.count())
    for ledUpdate in ledUpdate:
        rawquery = "select distinct l.id,l.nomor_kasus from led_ledlaporan lp join led_led l ON l.id = lp.led_laporan_id where l.nomor_kasus = " + "'" +str(ledUpdate.nomor_kasus) +"'"
        currLed = Led.objects.raw(rawquery)
        counter = 0
        for x in currLed:
            print(x.id)
            if(counter == 0):
                deepDiff(request,x.id)
                counter +=1
            else:
                break

#function untuk mendaptakan perubahan antarra data baru dan data lama
def deepDiff(request,currLed):
    #Get LED objects used in laporan
    laporanBaru = Led.objects.get(id = currLed)
    print("laporanbaru")
    print(laporanBaru.id)
    #Get all LED where consist of update date
    checkedLed = StateChange.objects.filter(nomor_kasus = laporanBaru).distinct('nomor_kasus2').values_list('nomor_kasus2')
    print(checkedLed)
    # for x in checkedLed:
    #     print(x.nomor_kasus2)
    
    #belum ditambahkan filter by tanggal update untuk memperkecil data,karena led yang memeiliki tanggal update merupakan led update ormis
    laporanBaru2 = Led.objects.filter(nomor_kasus = laporanBaru.nomor_kasus).exclude(id__in = [laporanBaru.id]).exclude(id__in = checkedLed)
    print("perbandingan2")
    print(laporanBaru2.count())
    
    stateDict= {'nomor_kasus':laporanBaru.id,
                'nomor_kasus2':"",
                'nama_kolom':"",
                'value':"",
                'status':""}
    ledDict = {
            "nomor_kasus": laporanBaru.nomor_kasus,
            "jumlah_kasus": laporanBaru.jumlah_kasus,
            "nama_pembuat": laporanBaru.nama_pembuat,
            "status": laporanBaru.status,
            "kode_cabang": laporanBaru.kode_cabang.kode_cabang,
            "kode_produk": laporanBaru.kode_produk.kode_produk,
            "nama_penemu": laporanBaru.nama_penemu,
            "kode_ktg_kejadian": laporanBaru.kode_ktg_kejadian.kode_kategori1,
            "kode_penyebab_level_1": laporanBaru.kode_penyebab.kode_penyebab1,
            "bussines_line": laporanBaru.business_line.kode_bisnis,
            "mata_uang" : laporanBaru.mata_uang,
            "nilai_tukar": laporanBaru.nilai_tukar,
            "kerugian_potensial": laporanBaru.kerugian_potensial,
            "rra"       : laporanBaru.rra,
            "recovery": laporanBaru.recovery,
            "kerugian_aktual": laporanBaru.kerugian_aktual,
            "tanggal_review": laporanBaru.tanggal_review,
            "reviewer": laporanBaru.reviewer,
            "hasil_review": laporanBaru.hasil_review,
            "keterangan_review":laporanBaru.keterangan_review,
            "summary_kejadian": laporanBaru.summary_kejadian,
            "kronologi_kejadian": laporanBaru.kronologi_kejadian,
            "tindakan_unit_kerja": laporanBaru.tindakan_unit_kerja,
            "tindakan_perbaikan": laporanBaru.tindakan_perbaikan,
            "dibuat_oleh": laporanBaru.dibuat_oleh,
            "disetujui_oleh": laporanBaru.disetujui_oleh,
            "risiko_kredit": laporanBaru.risiko_kredit
        
    }
    # print(type(ledDict))
    for ledObj in laporanBaru2:
        ledDict2 = {
                "nomor_kasus": ledObj.nomor_kasus,
                "jumlah_kasus": ledObj.jumlah_kasus,
                "nama_pembuat": ledObj.nama_pembuat,
                "status": ledObj.status,
                "kode_cabang": ledObj.kode_cabang.kode_cabang,
                "kode_produk": ledObj.kode_produk.kode_produk,
                "nama_penemu": ledObj.nama_penemu,
                "kode_ktg_kejadian": ledObj.kode_ktg_kejadian.kode_kategori1,
                "kode_penyebab_level_1": ledObj.kode_penyebab.kode_penyebab1,
                "bussines_line": ledObj.business_line.kode_bisnis,           
                "mata_uang": ledObj.mata_uang,
                "nilai_tukar": ledObj.nilai_tukar,
                "kerugian_potensial": ledObj.kerugian_potensial,
                "rra": ledObj.rra,
                "recovery": ledObj.recovery,
                "kerugian_aktual": ledObj.kerugian_aktual,
                "tanggal_review": ledObj.tanggal_review,
                "reviewer": ledObj.reviewer,
                "hasil_review": ledObj.hasil_review,
                "keterangan_review":ledObj.keterangan_review,
                "summary_kejadian": ledObj.summary_kejadian,
                "kronologi_kejadian": ledObj.kronologi_kejadian,
                "tindakan_unit_kerja": ledObj.tindakan_unit_kerja,
                "tindakan_perbaikan": ledObj.tindakan_perbaikan,
                "dibuat_oleh": ledObj.dibuat_oleh,
                "disetujui_oleh": ledObj.disetujui_oleh,
                "risiko_kredit": ledObj.risiko_kredit
        }  
        # print(ledDict2)
        # print("s")
        dictUpt= {'nomor_kasus2':ledObj.id}
        stateDict.update(dictUpt)
        diff = DeepDiff(ledDict,ledDict2)
        print(diff)
        columnName = ''
        newValues = ''
        diff2 = diff['values_changed']
        for x in diff2:
            index1 = x.index("['")  
            index2 =x.index("']")
            #ambil nama kolom
            columnName = x[index1+2:index2]
            print (columnName)
            dictUpt= {'nama_kolom':columnName}
            stateDict.update(dictUpt)
            
            # create_state(x)
            for y in diff2[x]:
                # index1 = y.index(":")
                # index2 = len(y)
                dictUpt= {'value':diff2[x][y]}
                stateDict.update(dictUpt)
                print (diff2[x][y])
                print("y")
                if(y == "new_value"):
                    create_state(request,stateDict)

#menampilkan nomor kasus yang memeiliki perbedaan antara data baru dan data yang sedang diproses pada laporan
def reviewLed(request):
    data = StateChange.objects.all().distinct('nomor_kasus')
    context = {
        'data':data,
    }
    return render(request,'led/reviewLed/reviewLed.html',context)

def reviewLedId(request,ledId):
    currLed = Led.objects.get(id= ledId)
    insertNewLED(request,currLed.id)
    
    stateChange =StateChange.objects.filter(nomor_kasus = currLed) 
    print(stateChange.count())
    
    #get laporan periode that has relation with database
    rawQuery = ("SELECT DISTINCT  lpr.nama,lpr.id,l.id" 
            +   " from led_ledlaporan llp"
            +   " join led_led l on llp.led_laporan_id = l.id "
            +   " join led_laporan lpr on lpr.id = llp.laporan_id"
            +   " join led_statechange sc on sc.nomor_kasus_id = l.id"
            +   " where l.id = " + str(currLed.id) +"")
    
    lprLinked = Laporan.objects.raw(rawQuery)
    for x in lprLinked:
        print(x.nama)
    context = {
        'data':currLed,
        'stateChange':stateChange,
        'lprLinked':lprLinked,
    }
    return render(request,'led/reviewLed/reviewLedId.html',context)


def insertNewLED (request,ledId):
    ledObj = Led.objects.get(id = ledId)
    
    if(ledObj.tanggal_update == None):
        print(ledObj.tanggal_update)
    else:
        print(ledObj.tanggal_update)


def replace(request,ledId,ledId2):
    currLED = Led.objects.get(id = ledId)
    updateLed = Led.objects.get(id=ledId2)
    namaUser = request.user.first_name
    column =  request.GET.get('column')
    sc=  request.GET.get('sc')
    print(column)
    
    if(column == "nama_pembuat"):
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Nama Pembuat",
            'dataLama':str(currLED.nama_pembuat),
            'dataBaru':str(updateLed.nama_pembuat),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        }
        
        currLED.nama_pembuat = updateLed.nama_pembuat
        currLED.save()
        perubahanData(request,dictPerubahan)
        deleteStateChange(sc)
    elif(column == "jumlah_kasus" ):
        
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Jumlah Kasus",
            'dataLama':str(currLED.jumlah_kasus),
            'dataBaru':str(updateLed.jumlah_kasus),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        }
        currLED.jumlah_kasus = updateLed.jumlah_kasus
        currLED.save()
        perubahanData(request,dictPerubahan)  
        deleteStateChange(sc)
    # elif(column == "kode_cabang" ):
    #     dictPerubahan ={
    #         'noKasus':currLED,
    #         'namaKolom':"Kode Cabang",
    #         'dataLama':str(currLED.kode_cabang),
    #         'dataBaru':str(updateLed.),
    #         'status':'',
    #         'alasan':'',
    #         'reviewer':currLED.reviewer,
    #         'user':namaUser,
    #     }
    #     currLED.jumlah_kasus = updateLed.jumlah_kasus
    #     currLED.save()    
    #     perubahanData(request,dictPerubahan)
    elif(column == "kode_produk" ):   
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Kode Produk",
            'dataLama':str(currLED.kode_produk.nama_produk),
            'dataBaru':str(updateLed.kode_produk.nama_produk),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        }
        currLED.kode_produk = updateLed.kode_produk
          
        perubahanData(request,dictPerubahan)
        deleteStateChange(sc)
    elif(column == "nama_penemu" ):
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Jumlah Kasus",
            'dataLama':str(currLED.nama_penemu),
            'dataBaru':str(updateLed.nama_penemu),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        }
        currLED.nama_penemu = updateLed.nama_penemu
               
        perubahanData(request,dictPerubahan)
        deleteStateChange(sc)
    elif(column == "kode_ktg_kejadian" ):
        
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Kode Kategori",
            'dataLama':str(currLED.kode_ktg_kejadian.kode_kategori),
            'dataBaru':str(updateLed.kode__ktg_kejadian.kode_kategori),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        }
        currLED.kode_ktg_kejadian = updateLed.kode_ktg_kejadian
        
        perubahanData(request,dictPerubahan)
        deleteStateChange(sc)
    elif(column == "kode_penyebab" ): 
           
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Kode Penyebab",
            'dataLama':str(currLED.kode_ktg_kejadian.kode_kategori),
            'dataBaru':str(updateLed.kode__ktg_kejadian.kode_kategori),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        }       
        currLED.kode_penyebab= updateLed.kode_penyebab  
        
        perubahanData(request,dictPerubahan)
        deleteStateChange(sc)
    elif(column == "bussines_line"):
        
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Bussines Line",
            'dataLama':str(currLED.business_line.kode_bisnis),
            'dataBaru':str(updateLed.business_line.kode_bisnis),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        }   
        currLED.busnes_line = updateLed.business_line
        
        deleteStateChange(sc)
    elif(column == "coa_biaya" ):
        
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Coa Biaya",
            'dataLama':str(currLED.coa_biaya.nomor_kasus),
            'dataBaru':str(updateLed.coa_biaya.nomor_kasus),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        }  
        currLED.coa_biaya = updateLed.coa_biaya
        
        perubahanData(request,dictPerubahan)
        deleteStateChange(sc)
    elif(column == "mata_uang" ):
        
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Mata Uang",
            'dataLama':str(currLED.mata_uang),
            'dataBaru':str(updateLed.mata_uang),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        } 
        currLED.mata_uang = updateLed.mata_uang  
        
        perubahanData(request,dictPerubahan)
        deleteStateChange(sc)
    elif(column == "nilai_tukar" ):
        
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Nilai Tukar",
            'dataLama':str(currLED.nilai_tukar),
            'dataBaru':str(updateLed.nilai_tukar),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        }  
        currLED.nilai_tukar = updateLed.nilai_tukar
       
        perubahanData(request,dictPerubahan)
        deleteStateChange(sc)
    elif(column == "kerugian_potensial" ):
        
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Kerugian Potensial",
            'dataLama':str(currLED.kerugian_potensial),
            'dataBaru':str(updateLed.kerugian_ptensial),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        } 
        
       
        perubahanData(request,dictPerubahan)
        deleteStateChange(sc)
    elif(column == "rra" ):
        
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"RRA",
            'dataLama':str(currLED.rra),
            'dataBaru':str(updateLed.rra),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        } 
        currLED.rra = updateLed.rra
        
        perubahanData(request,dictPerubahan)
        deleteStateChange(sc)
    elif(column == "recovery" ):
        currLED.recovery = updateLed.recovery
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Recovery",
            'dataLama':str(currLED.recovery),
            'dataBaru':str(updateLed.recovery),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        } 
        
        perubahanData(request,dictPerubahan)
        deleteStateChange(sc)
    
    elif(column == "kerugian_aktual" ):
        
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Kerugian Aktual",
            'dataLama':str(currLED.kerugian_aktual),
            'dataBaru':str(updateLed.kerugian_aktual),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        } 
        currLED.kerugian_aktual = updateLed.kerugian_aktual
        
        perubahanData(request,dictPerubahan)
        deleteStateChange(sc)
    elif(column == "tanggal_review" ):
        
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Tanggal Review",
            'dataLama':str(currLED.tanggal_review),
            'dataBaru':str(updateLed.tanggal_review),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        } 
        currLED.jumlah_kasus = updateLed.jumlah_kasus
        
        perubahanData(request,dictPerubahan)
        deleteStateChange(sc)
    elif(column == "reviewer" ):
        currLED.reviewer = updateLed.reviewer
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Reviewer",
            'dataLama':str(currLED.reviewer),
            'dataBaru':str(updateLed.reviewer),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        } 
        
        perubahanData(request,dictPerubahan)
        deleteStateChange(sc)
    elif(column == "hasil_review" ):
        currLED.hasil_review = updateLed.hasil_review
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Hasil Review",
            'dataLama':str(currLED.hasil_review),
            'dataBaru':str(updateLed.hasil_review),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        } 
        
        perubahanData(request,dictPerubahan)
        deleteStateChange(sc)
    elif(column == "keterangan_review" ):
        currLED.keterangan_review = updateLed.keterangan_review
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Kerugian Potensial",
            'dataLama':str(currLED.keterangan_review),
            'dataBaru':str(updateLed.keteranga_review),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        } 
        
        perubahanData(request,dictPerubahan)
        deleteStateChange(sc)
    elif(column == "summary_kejadian" ):
        currLED.summary_kejadian = updateLed.summary_kejadian
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Summary Kejadian",
            'dataLama':str(currLED.summary_kejadian),
            'dataBaru':str(updateLed.summary_kejadian),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        } 
        
        
        perubahanData(request,dictPerubahan)
        deleteStateChange(sc)
        
    elif(column == "kronologi_kejadian" ):
        currLED.kronologi_kejadian = updateLed.kronologi_kejadian
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Kronologi Kejadian",
            'dataLama':str(currLED.kronologi_kejadian),
            'dataBaru':str(updateLed.kronologi_kejadian),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        } 
        
        perubahanData(request,dictPerubahan)
        deleteStateChange(sc)
    elif(column == "tindakan_unit_kerja" ):
        currLED.tindakan_unit_kerja = updateLed.tindakan_unit_kerja
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Tindakan Unit Kerja",
            'dataLama':str(currLED.tindakan_unit_kerja),
            'dataBaru':str(updateLed.tindakan_unit_kerja),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        } 
        
       
        perubahanData(request,dictPerubahan)
        deleteStateChange(sc)
    elif(column == "tindakan_perbaikan" ):
        currLED.tindakan_perbaikan = updateLed.tindakan_perbaikan
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Tindakan Perbaikan",
            'dataLama':str(currLED.tindakan_perbaikan),
            'dataBaru':str(updateLed.tindakan_perbaikan),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        } 
        
        perubahanData(request,dictPerubahan)
        deleteStateChange(ledId2)
    elif(column == "dibuat_oleh" ):
        currLED.dibuat_oleh = updateLed.dibuat_oleh
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Dibuat Oleh",
            'dataLama':str(currLED.dibuat_oleh),
            'dataBaru':str(updateLed.dibuat_oleh),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        } 
        
        perubahanData(request,dictPerubahan)
        deleteStateChange(sc)
    elif(column == "disetujui_oleh" ):
        currLED.disetujui_oleh = updateLed.disetujui_oleh
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Kerugian Potensial",
            'dataLama':str(currLED.disetujui_oleh),
            'dataBaru':str(updateLed.disetujui_oleh),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        } 
        
        perubahanData(request,dictPerubahan)
        deleteStateChange(sc)
    elif(column == "risiko_kredit" ):
        currLED.risiko_kredit = updateLed.risiko_kredit
        dictPerubahan ={
            'noKasus':currLED,
            'namaKolom':"Kerugian Potensial",
            'dataLama':str(currLED.disetujui_oleh),
            'dataBaru':str(updateLed.disetujui_oleh),
            'status':'',
            'alasan':'',
            'reviewer':currLED.reviewer,
            'user':namaUser,
        } 
       
        perubahanData(request,dictPerubahan)
        deleteStateChange(sc)
    return redirect('reviewLedId',ledId = ledId)

def editLed(request,ledId):
    currLed = Led.objects.get(id =ledId)
    kd_cabang = Cabang.objects.all()
    kd_unit = Unit.objects.all()
    kd_produk = Produk.objects.all()
    kt_kejadian = KategoriKejadian.objects.all()
    ls_penyebab = Penyebab.objects.all()
    ls_bl   = BusinessLine.objects.all()
    ls_coa = Coa.objects.all()
    ls_journal = Journal.objects.filter(nomor_kasus = currLed.nomor_kasus)
    print(ls_journal.count())
    
    print(currLed.tanggal_kejadian)
    context={
        'currLed':currLed,
        'kd_cabang':kd_cabang,
        'kd_produk':kd_produk,
        'kd_unit':kd_unit,
        'kt_kejadian':kt_kejadian,
        'ls_penyebab':ls_penyebab,
        'ls_bl':ls_bl,
        'ls_coa':ls_coa,
        'ls_journal':ls_journal,
        'jj':json.dumps("ss"),
    }
    return render(request,'led/editLed/editLed.html',context)
#Kurang
#validate spasi dara traling space untuk menghindari perubahan data hanya karena space
#belum ditambah user
def updateLed(request,ledId):
    print("editLed")
    if request.POST:
        print("post")
        
        nomorKasus = request.POST.get('no_kasus','None')
        dibuat_oleh = request.POST.get('dibuat_oleh','None')
        tgl_input    = request.POST.get('tgl_input','None')
        tipe_led  = request.POST.get('tipe_kasus','None')
        kode_kanwil = request.POST.get('kode_kanwil','None')
        kode_kcu = request.POST.get('kode_kcu','None')
        kode_unit = request.POST.get('kode_unit',None)
        tgl_kejadian = request.POST.get('kode_unit','None')
        kode_produk = request.POST.get('kode_produk','None')
        nama_produk = request.POST.get('nama_produk','None')
        tgl_identifikasi = request.POST.get('tgl_identifikasi','None')
        ditemukan_oleh = request.POST.get('ditemukan_oleh','None')
        bussines_line = request.POST.get('business_line','None')
        kode_kejadian = request.POST.get('kode_kejadian','None')
        kode_penyebab = request.POST.get('kode_penyebab','None')
        summary_kejadian = request.POST.get('summary_kejadian','None')
        kronologi_kejadian = request.POST.get('kronologi_kejadian','None')
        tindakan_unit_kerja = request.POST.get('tindakan_unit_kerja','None')
        mata_uang = request.POST.get('mata_uang','None')
        nilai_tukar = request.POST.get('nilai_tukar','None')
        kerugian_potensial = request.POST.get('kerugian_potensial','None')
        kode_rra = request.POST.get('kode_rra','None')
        recovery = request.POST.get('recovery','None')
        kerugian_aktual = request.POST.get('kerugian_aktual','None')
        hasil_review = request.POST.get('hasil_review','None')
        keterangan_review = request.POST.get('keterangan_review','None')
        status = request.POST.get('status','None')
        tipe_kasus = request.POST.get('tipe_kasus','None')
        jml_kasus = request.POST.get('jml_kasus','None')
        risiko_kredit = request.POST.get('risiko_kredit','None')
        tanggal_selesai = request.POST.get('tanggal_selesai','None')
        tanggal_update = request.POST.get('tanggal_update','None')
        tanggal_review = request.POST.get('tanggal_review','None')
        disetujui_oleh = request.POST.get('disetujui_oleh','None')
        ditemukan_oleh = request.POST.get('ditemukan_oleh','None')
        reviewer = request.POST.get('reviewer','None')
        tindakan_unit_kerja = request.POST.get('tindakan_unit_kerja','None')
        tindakan_perbaikan = request.POST.get('tindakan_perbaikan','None')
       
        
        namaUser = request.user.first_name
        #get led
        currLed = Led.objects.get(id = ledId)
        print(dict(request.POST.items()))
        
        dictPerubahan ={
            'noKasus':currLed,
            'namaKolom':'',
            'dataLama':'',
            'dataBaru':'',
            'status':'',
            'alasan':'',
            'reviewer':'',
            'user':'',
        }
        
        # if(currLed.tanggal_kejadian != tgl_kejadian and tgl_kejadian!= 'None'):
        #     currLed.tanggal_kejadian = tgl_kejadian
        # if(currLed.tanggal_teridentifikasi != tgl_identifikasi  and tgl_identifikasi != 'None'):
        #     currLed.tanggal_teridentifikasi = tgl_identifikasi
        # if(currLed.tanggal_input != tgl_input  and tgl_input!= 'None'):
        #     currLed.tanggal_input = tgl_input
        # if(currLed.tanggal_selesai != tanggal_selesai  and tanggal_selesai!= 'None'):
        #     currLed.tanggal_selesai = tanggal_selesai
        # if( currLed.tanggal_update != tanggal_update  and tanggal_update != 'None' ):
        #     currLed.tanggal_update = tanggal_update
        # if(str(currLed.jumlah_kasus) != jml_kasus and jml_kasus != 'None'):
          
        #     dictPerubahan ={
        #     'noKasus':currLed,
        #     'namaKolom':"Jumlah Kasus",
        #     'dataLama':str(currLed.jumlah_kasus),
        #     'dataBaru':str(jml_kasus),
        #     'status':'',
        #     'alasan':'',
        #     'reviewer':currLed.reviewer,
        #     'user':namaUser,
        # }
        #     currLed.jumlah_kasus =jml_kasus 
        #     perubahanData(request,dictPerubahan)    
        # if(str(currLed.nama_pembuat) != dibuat_oleh):
           
            
        #     dictPerubahan ={
        #     'noKasus':currLed,
        #     'namaKolom':"Nama Pembuat",
        #     'dataLama':str(currLed.nama_pembuat),
        #     'dataBaru':str(dibuat_oleh),
        #     'status':'',
        #     'alasan':'',
        #     'reviewer':currLed.reviewer,
        #     'user':namaUser,
        # }
        #     currLed.nama_pembuat = dibuat_oleh
        #     perubahanData(request,dictPerubahan)  
        # if( currLed.status != status ):
        #     dictPerubahan ={
        #     'noKasus':currLed,
        #     'namaKolom':"status",
        #     'dataLama':str(currLed.status),
        #     'dataBaru':str(status),
        #     'status':'',
        #     'alasan':'',
        #     'reviewer':currLed.reviewer,
        #     'user':namaUser,
        # }
        #     currLed.status = status
        #     perubahanData(request,dictPerubahan)  
            
        # if( currLed.kode_unit !=  kode_unit and kode_unit != 'None'):
        #     dictPerubahan ={
        #     'noKasus':currLed,
        #     'namaKolom':"Kode Unit",
        #     'dataLama':str(currLed.kode_unit),
        #     'dataBaru':str(kode_unit),
        #     'status':'',
        #     'alasan':'',
        #     'reviewer':currLed.reviewer,
        #     'user':namaUser,
        # }
        #     currLed.kode_unit = kode_unit
        #     perubahanData(request,dictPerubahan)  
        # if( str(currLed.tipe_led) != tipe_led  and tipe_led != 'None'):
        #     dictPerubahan ={
        #     'noKasus':currLed,
        #     'namaKolom':"Tipe LED",
        #     'dataLama':str(currLed.tipe_led),
        #     'dataBaru':str(tipe_led),
        #     'status':'',
        #     'alasan':'',
        #     'reviewer':currLed.reviewer,
        #     'user':namaUser,
        # }
        #     currLed.tipe_led == tipe_led
        #     perubahanData(request,dictPerubahan)  
        # if(str(currLed.kode_produk.id) != nama_produk):
        #     print(nama_produk)
        #     print(currLed.kode_produk.id)
        #     print(len(currLed.kode_produk.nama_produk))
        #     print(len(nama_produk))
        #     produk = Produk.objects.get(id = nama_produk)
        #     dictPerubahan ={
        #     'noKasus':currLed,
        #     'namaKolom':"Produk",
        #     'dataLama':str(currLed.produk),
        #     'dataBaru':str(produk.nama_produk),
        #     'status':'',
        #     'alasan':'',
        #     'reviewer':currLed.reviewer,
        #     'user':namaUser,
        # }
        #     currLed.kode_produk= produk.nama
        #     perubahanData(request,dictPerubahan)  
        # if(currLed.nama_penemu != ditemukan_oleh):
        #     currLed.nama_penemu= ditemukan_oleh
        # if(str(currLed.kode_ktg_kejadian.id) != str(kode_kejadian)):
        #     print(currLed.kode_ktg_kejadian)
        #     print(kode_kejadian)
        #     kodeKejadian = KategoriKejadian.objects.get(id= kode_kejadian)
        #     print(kodeKejadian.kode_kategori1)
        #     dictPerubahan ={
        #     'noKasus':currLed,
        #     'namaKolom':"Kategori Kejadian",
        #     'dataLama':str(currLed.kode_ktg_kejadian),
        #     'dataBaru':str(kodeKejadian),
        #     'status':'',
        #     'alasan':'',
        #     'reviewer':currLed.reviewer,
        #     'user':namaUser,
        #     }
        #     perubahanData(request,dictPerubahan)  
        #     currLed.kode_ktg_kejadian = kodeKejadian      
        # if(str(currLed.kode_penyebab.id) != kode_penyebab):
        #     kodePenyebab = Penyebab.objects.get(id = kode_penyebab)
        #     dictPerubahan ={
        #     'noKasus':currLed,
        #     'namaKolom':"Kode Penyebab",
        #     'dataLama':str(currLed.kode_penyebab),
        #     'dataBaru':str(kode_penyebab),
        #     'status':'',
        #     'alasan':'',
        #     'reviewer':currLed.reviewer,
        #     'user':namaUser,
        #     }
        #     currLed.kode_penyebab = kodePenyebab
        #     perubahanData(request,dictPerubahan)  
            
        # if(str(currLed.kode_produk.id) != kode_produk and kode_produk != 'None'):
        #     kodeProduk = Produk.objects.get(id = kode_produk)
        #     dictPerubahan ={
        #     'noKasus':currLed,
        #     'namaKolom':"Kode Produk",
        #     'dataLama':str(currLed.kode_produk),
        #     'dataBaru':str(kode_produk),
        #     'status':'',
        #     'alasan':'',
        #     'reviewer':currLed.reviewer,
        #     'user': namaUser,
        #     }
        #     currLed.kode_produk = kodeProduk
        #     perubahanData(request,dictPerubahan)  
        # if(bussines_line != ''):
        #     print("business_line "+ bussines_line)
        #     if(str(currLed.business_line.id) != bussines_line and bussines_line !='None'):
                
        #         bussinesLine = BusinessLine.objects.get(id = bussines_line)
        #         dictPerubahan ={
        #         'noKasus':currLed,
        #         'namaKolom':"Business Line",
        #         'dataLama':str(currLed.business_line),
        #         'dataBaru':str(bussines_line),
        #         'status':'',
        #         'alasan':'',
        #         'reviewer':currLed.reviewer,
        #         'user':namaUser,
        #     }
        #     currLed.business_line = bussinesLine
        #     perubahanData(request,dictPerubahan)  
        # if(currLed.mata_uang !=  mata_uang):
        #     dictPerubahan ={
        #     'noKasus':currLed,
        #     'namaKolom':"Mata Uang",
        #     'dataLama':str(currLed.mata_uang),
        #     'dataBaru':str(mata_uang),
        #     'status':'',
        #     'alasan':'',
        #     'reviewer':currLed.reviewer,
        #     'user':namaUser,
        # }
        #     currLed.mata_uang =  mata_uang
        #     perubahanData(request,dictPerubahan)  
        # if(str(currLed.nilai_tukar) != nilai_tukar):
        #     dictPerubahan ={
        #     'noKasus':currLed,
        #     'namaKolom':"Nilai Tukar",
        #     'dataLama':str(currLed.nilai_tukar),
        #     'dataBaru':str(nilai_tukar),
        #     'status':'',
        #     'alasan':'',
        #     'reviewer':currLed.reviewer,
        #     'user':namaUser,
        # }
        #     currLed.nilai_tukar = nilai_tukar
        #     perubahanData(request,dictPerubahan)  
        # if(str(currLed.kerugian_potensial) != kerugian_potensial):
        #     dictPerubahan ={
        #     'noKasus':currLed,
        #     'namaKolom':"kerugian Potensial",
        #     'dataLama':str(currLed.kerugian_potensial),
        #     'dataBaru':str(kerugian_potensial),
        #     'status':'',
        #     'alasan':'',
        #     'reviewer':currLed.reviewer,
        #     'user':namaUser,
        # }
        #     currLed.kerugian_potensial = kerugian_potensial
        #     perubahanData(request,dictPerubahan)  
        # if(str(currLed.rra) != kode_rra):
          
        #     dictPerubahan ={
        #     'noKasus':currLed,
        #     'namaKolom':"RRA",
        #     'dataLama':str(currLed.rra),
        #     'dataBaru':str(kode_rra),
        #     'status':'',
        #     'alasan':'',
        #     'reviewer':currLed.reviewer,
        #     'user':namaUser,
        # }
        #     print('rra')
        #     print(type(kode_rra))
        #     currLed.rra = kode_rra
        #     perubahanData(request,dictPerubahan)  
        # if(str(currLed.recovery) != recovery):
        #     dictPerubahan ={
        #     'noKasus':currLed,
        #     'namaKolom':"Recovery",
        #     'dataLama':str(currLed.recovery),
        #     'dataBaru':str(recovery),
        #     'status':'',
        #     'alasan':'',
        #     'reviewer':currLed.reviewer,
        #     'user':namaUser,
        # }
        #     currLed.recovery = recovery
        #     perubahanData(request,dictPerubahan)  
        # if(str(currLed.kerugian_aktual) != kerugian_aktual):
        #     dictPerubahan ={
        #     'noKasus':currLed,
        #     'namaKolom':"Kerugian Aktual",
        #     'dataLama':str(currLed.kerugian_aktual),
        #     'dataBaru':str(kerugian_aktual),
        #     'status':'',
        #     'alasan':'',
        #     'reviewer':currLed.reviewer,
        #     'user':namaUser,
        # }
        #     currLed.kerugian_aktual = kerugian_aktual
        #     perubahanData(request,dictPerubahan)  
        # if(currLed.tanggal_review != tanggal_review  and tanggal_review!= 'None'):
        #     dictPerubahan ={
        #     'noKasus':currLed,
        #     'namaKolom':"Tanggal Review",
        #     'dataLama':str(currLed.tanggal_review),
        #     'dataBaru':str(tanggal_review),
        #     'status':'',
        #     'alasan':'',
        #     'reviewer':currLed.reviewer,
        #     'user':namaUser,
        # }
        #     currLed.tanggal_review = tanggal_review
        #     perubahanData(request,dictPerubahan)  
        # if(currLed.reviewer != reviewer):
        #     dictPerubahan ={
        #     'noKasus':currLed,
        #     'namaKolom':"Reviewer",
        #     'dataLama':str(currLed.reviewer),
        #     'dataBaru':str(reviewer),
        #     'status':'',
        #     'alasan':'',
        #     'reviewer':currLed.reviewer,
        #     'user':namaUser,
        #     }
        #     currLed.reviewer = reviewer
        # # if(currLed.tipe_kasus != tipe_kasus and tipe_kasus != 'None'):
        # #     dictPerubahan ={
        # #     'noKasus':currLed,
        # #     'namaKolom':"Tipe Kasus",
        # #     'dataLama':str(currLed.tipe_kasus),
        # #     'dataBaru':str(reviewer.tipe_kasus),
        # #     'status':'',
        # #     'alasan':'',
        # #     'reviewer':currLed.reviewer,
        # #     'user':'',
        # #     }
        # #     currLed.tipe_kasus = tipe_kasus
        # #     perubahanData(request,dictPerubahan)
            
        # if(currLed.hasil_review != hasil_review):
        #     currLed.hasil_review = hasil_review
        # if(currLed.keterangan_review != keterangan_review):
        #     currLed.keterangan_review = keterangan_review
        # if(currLed.summary_kejadian != summary_kejadian):
        #     currLed.summary_kejadian = summary_kejadian
        # if(currLed.kronologi_kejadian != kronologi_kejadian):
        #     currLed.kronologi_kejadian = kronologi_kejadian
        # if(currLed.tindakan_unit_kerja != tindakan_unit_kerja):
        #     currLed.tindakan_unit_kerja = tindakan_unit_kerja
        # if(currLed.tindakan_perbaikan != tindakan_perbaikan):
        #     currLed.tindakan_perbaikan = tindakan_perbaikan
        # if(currLed.dibuat_oleh != dibuat_oleh):
        #     currLed.dibuat_oleh = dibuat_oleh
        # if(currLed.disetujui_oleh != disetujui_oleh):
        #     currLed.disetujui_oleh = disetujui_oleh
        # if(currLed.ditemukan_oleh != ditemukan_oleh and ditemukan_oleh !='None'):
        #     currLed.ditemukan_oleh = ditemukan_oleh
        # if(currLed.risiko_kredit != risiko_kredit):
        #     currLed.risiko_kredit = risiko_kredit
        
        #Update review user
        review_user = request.POST.get('btn','')
        if(review_user == 'Submit'):
            
            dictPerubahan ={
            'noKasus':currLed,
            'namaKolom':"review_user",
            'dataLama':str(currLed.review_user),
            'dataBaru':str("Submit"),
            'status':'',
            'alasan':'',
            'reviewer':currLed.reviewer,
            'user':namaUser,
        }
            currLed.review_user = "Submit"
            currLed.save()
            perubahanData(dictPerubahan) 
        else:
            dictPerubahan ={
            'noKasus':currLed,
            'namaKolom':"review_user",
            'dataLama':str(currLed.review_user),
            'dataBaru':str("Save"),
            'status':'',
            'alasan':'',
            'reviewer':currLed.reviewer,
            'user':namaUser,
            }
            currLed.review_user = "Save"
            currLed.save()
            perubahanData(dictPerubahan)
        # currLed.save()
        print('ss')
       
    
        return editLed(request,ledId)

def perubahanData(dictPerubahan):
    # 'noKasus':currLed,
    #         'namaKolom':"Jumlah Kasus",
    #         'dataLama':str(currLed.jumlah_kasus),
    #         'dataBaru':str(jml_kasus),
    #         'status':'',
    #         'alasan':'',
    #         'reviewer':currLed.reviewer,
    #         'user':'',
    print("gg")
    perubahan = Perubahan()
    perubahan.nomor_kasus = dictPerubahan["noKasus"]
    perubahan.nama_kolom = dictPerubahan["namaKolom"]
    perubahan.data_lama = dictPerubahan["dataLama"]
    perubahan.data_baru = dictPerubahan["dataBaru"]
    perubahan.status = dictPerubahan["status"]
    perubahan.alasan= dictPerubahan["alasan"]
    perubahan.user =dictPerubahan["user"]
    perubahan.save()
    
    return True
    # for key,value in kwargs.items():
    #     # print("%s , %s" %(key,value))

#list -list kategori
def listCabang(request):
    kd_cabang = Cabang.objects.all()
    context={
        'kd_cabang':kd_cabang
        }
    return render(request,'led/list/cabang.html',context)

def listProduk(request):
    kd_produk = Produk.objects.all()
    context={
        'kd_produk':kd_produk
        }
    return render(request,'led/list/produk.html',context)

def listKejadian(request):
    kt_kejadian = KategoriKejadian.objects.all().order_by('kode_kategori1','nama_kategori1')
    context={
        'kt_kejadian':kt_kejadian
        }
    return render(request,'led/list/kategoriKejadian.html',context)

def listPenyebab(request):
    ls_penyebab = Penyebab.objects.all()
    
    
    context={
        'ls_penyebab':ls_penyebab
        }
    return render(request,'led/list/penyebab.html',context)

def listBusinessLine(request):
    ls_bl = BusinessLine.objects.all()
    context={
        'ls_bl':ls_bl
        }
    return render(request,'led/list/businessLine.html',context)

def listCoa(request):
    ls_coa = Coa.objects.all()
    context={
        'ls_coa':ls_coa
        }
    return render(request,'led/list/coa.html',context)

def exportLaporan(request,ledId):
    laporan = Laporan.objects.get(id = ledId)

    #decide file name
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment;' + 'filename=' +  str(laporan.nama)+'.xlsx'
    
    x = "Risiko Operasional"
    ledLaporan= Led.objects.raw("Select l.id,l.nomor_kasus,l.tanggal_kejadian,l.tanggal_input,l.tanggal_teridentifikasi,l.tanggal_input,l.tanggal_selesai,l.jumlah_kasus,"
                                +"l.nama_pembuat,l.status,l.kode_cabang_id,l.kode_produk_id,l.nama_penemu,l.kode_ktg_kejadian_id,l.kode_penyebab_id,l.business_line_id,l.coa_biaya_id,l.mata_uang,l.nilai_tukar,l.kerugian_potensial,l.rra,"
                                +"l.recovery,l.kerugian_aktual,l.tanggal_review,l.reviewer,l.hasil_review,l.keterangan_review,l.summary_kejadian,l.kronologi_kejadian,"
                                +"l.tindakan_unit_kerja,l.tindakan_perbaikan,l.dibuat_oleh,l.disetujui_oleh,l.risiko_kredit"
                                +" From led_led l join led_ledlaporan ll on l.id = ll.led_laporan_id join led_laporan lap on lap.id = ll.laporan_id  where lap.id = %s",[ledId])
    #takenRisiko = IdentifikasiRisiko.objects.raw("SELECT DISTINCT IR.id,LENGTH(idn_komentar) AS panjang_komentar,RA.inherent,idn_deskripsi,idn_likelihood,idn_dampak,idn_nilai_risiko,identifikasi_risiko_id,mtg_deskripsi,mtg_likelihood,mtg_dampak,mtg_nilai_risiko,mtg_keterangan_likelihood,mtg_keterangan_dampak FROM public.library_riskassessment RA JOIN library_identifikasirisiko IR ON RA.id = IR.idn_risk_assessment_id JOIN library_mitigasirisiko MR ON MR.identifikasi_risiko_id = IR.id JOIN library_tiperisiko tp ON IR.tipe_risiko_id = tp.id WHERE RA.id = %s AND tp.tipe_risiko_nama LIKE 'Risiko Hukum' ",[raID])
    print(LedLaporan)
    wb = Workbook()
    ws = wb.active
    
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment;' + 'filename=' +  str(laporan.nama)+'.xlsx'
    #create heading
    cell = ws.cell(row=1,column=1)
    cell.value = 'Id.'
    cell.alignment = Alignment(horizontal='center', vertical='center')  
  
    cell = ws.cell(row=1,column=2)
    cell.value = 'Tanggal Kejadian'
    cell.alignment = Alignment(horizontal='center', vertical='center')

    cell = ws.cell(row=1,column=3)
    cell.value = 'Tanggal Input'
    cell.alignment = Alignment(horizontal='center', vertical='center')  

    cell = ws.cell(row=1,column=4)
    cell.value = 'Tanggal Terindentifikasi'
    cell.alignment = Alignment(horizontal='center', vertical='center') 
    
    cell = ws.cell(row=1,column=5)
    cell.value = 'Tanggal Input'
    cell.alignment = Alignment(horizontal='center', vertical='center')  
    
    cell = ws.cell(row=1,column=6)
    cell.value = 'Tanggal Selesai'
    cell.alignment = Alignment(horizontal='center', vertical='center') 
    
    cell = ws.cell(row=1,column=7)
    cell.value = 'Tanggal Update'
    cell.alignment = Alignment(horizontal='center', vertical='center')  
    
    cell = ws.cell(row=1,column=8)
    cell.value = 'Jumlah Kasus'
    cell.alignment = Alignment(horizontal='center', vertical='center')    
    
    cell = ws.cell(row=1,column=9)
    cell.value = 'Tipe LED'
    cell.alignment = Alignment(horizontal='center', vertical='center')   

    cell = ws.cell(row=1,column=10)
    cell.value = 'Status'
    cell.alignment = Alignment(horizontal='center', vertical='center')   


    cell = ws.cell(row=1,column=11)
    cell.value = 'Kode Cabang'
    cell.alignment = Alignment(horizontal='center', vertical='center')  
    
    cell = ws.cell(row=1,column=12)
    cell.value = 'Nama Cabang'
    cell.alignment = Alignment(horizontal='center', vertical='center')    
    
    cell = ws.cell(row=1,column=13)
    cell.value = 'Kode Unit Kerja'
    cell.alignment = Alignment(horizontal='center', vertical='center')  
    
    cell = ws.cell(row=1,column=14)
    cell.value = 'Nama Unit Kerja'
    cell.alignment = Alignment(horizontal='center', vertical='center')    

    cell = ws.cell(row=1,column=15)
    cell.value = 'Nama Produk'
    cell.alignment = Alignment(horizontal='center', vertical='center')   

    cell = ws.cell(row=1,column=16)
    cell.value = 'Ditemukan Oleh'
    cell.alignment = Alignment(horizontal='center', vertical='center')  
    
    cell = ws.cell(row=1,column=17)
    cell.value = 'Kode Kategori Kejadian'
    cell.alignment = Alignment(horizontal='center', vertical='center')    
    
    cell = ws.cell(row=1,column=18)
    cell.value = 'Kode Penyebab'
    cell.alignment = Alignment(horizontal='center', vertical='center')  
    
    cell = ws.cell(row=1,column=19)
    cell.value = 'Business Line'
    cell.alignment = Alignment(horizontal='center', vertical='center')    
    
    cell = ws.cell(row=1,column=20)
    cell.value = 'Coa Biaya'
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    cell = ws.cell(row=1,column=21)
    cell.value = 'Mata Uang'
    cell.alignment = Alignment(horizontal='center', vertical='center')   

    cell = ws.cell(row=1,column=22)
    cell.value = 'Nilai Tukar'
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    cell = ws.cell(row=1,column=23)
    cell.value = 'Kerugian potensial'
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    cell = ws.cell(row=1,column=24)
    cell.value = 'RRA'
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    cell = ws.cell(row=1,column=25)
    cell.value = 'Recovery'
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    cell = ws.cell(row=1,column=26)
    cell.value = 'Kerugian Aktual'
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    cell = ws.cell(row=1,column=27)
    cell.value = 'Tanggal Review'
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    cell = ws.cell(row=1,column=28)
    cell.value = 'Reviewer'
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    cell = ws.cell(row=1,column=29)
    cell.value = 'Hasil Review'
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    cell = ws.cell(row=1,column=30)
    cell.value = 'Keterangan Review'
    cell.alignment = Alignment(horizontal='center', vertical='center') 
    
    cell = ws.cell(row=1,column=31)
    cell.value = 'Summary Kejadian'
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    cell = ws.cell(row=1,column=32)
    cell.value = 'Kronologo Kejadian'
    cell.alignment = Alignment(horizontal='center', vertical='center')     
    
    cell = ws.cell(row=1,column=33)
    cell.value = 'Tindakan Unit Kerja'
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    

    #looping from row 1
    idrow= 2
    idcol = 1

   
    for data in ledLaporan:       
        #deskripsi identifikasi
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.id
        idcol+=1
   
        #likelihood
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.nomor_kasus
        idcol+=1
        #dampak
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = "data.tanggal_kejadian"
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = "data.tanggal_teridentifikasi"
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value =" data.tanggal_input"
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = "data.tanggal_update"
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.jumlah_kasus
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = ""
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.status
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = str(data.kode_cabang)
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = "data.nama_cabang"
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = ""
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = ""
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = ""
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = ""
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.nama_penemu
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = str(data.kode_ktg_kejadian)
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = str(data.kode_penyebab)
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value =str(data.business_line)
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = str(data.coa_biaya)
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.mata_uang
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.nilai_tukar
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.kerugian_potensial
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.rra
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.recovery
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.kerugian_aktual
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.tanggal_review
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.keterangan_review
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.summary_kejadian
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.kronologi_kejadian
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.tindakan_unit_kerja
        idcol+=1
        
        idrow += 1
        idcol =1
    
    
    print(idrow) 
           
      
    wb.save(response)

    wb.close()
    context = {
    }
    return response

def xlsToXlsx(request):
    pwd = os.path.dirname(__file__)
    print(pwd)
    for filename in os.listdir('.\led\static\\'):
        print(filename)
        if filename.endswith('.XLS'):
            # df = pd.read_excel(filename)
            print("xls")
            print(type(filename))
            destName = filename.replace('.XLS','.xlsx')
            #create file baru dengan format xlsx
            px.save_book_as(file_name= pwd + '\static\\'+ filename, dest_file_name= pwd + '\static\\'+destName )
            #delete file lama/#ganti nama file lama
            # os.remove(pwd + '\static\\'+ filename)
    
def iptKtgKejadian(request):
    pwd = os.path.dirname(__file__)
    file = pwd + '\static\periode.xlsx'
    file = load_workbook(file)

    

    #read file
    
    #Looping row
    #insert data
    #delete
    #return
    
def compImport (request):
    
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment;' + 'filename=' +  "s"+'.xlsx'
    
    ledLaporan= Led.objects.raw("Select l.id,l.nomor_kasus,l.tanggal_kejadian,l.tanggal_input,l.tanggal_teridentifikasi,l.tanggal_input,l.tanggal_selesai,l.jumlah_kasus,"
                                +"l.nama_pembuat,l.status,l.kode_cabang_id,l.kode_produk_id,l.nama_penemu,l.kode_ktg_kejadian_id,l.kode_penyebab_id,l.business_line_id,l.coa_biaya_id,l.mata_uang,l.nilai_tukar,l.kerugian_potensial,l.rra,"
                                +"l.recovery,l.kerugian_aktual,l.tanggal_review,l.reviewer,l.hasil_review,l.keterangan_review,l.summary_kejadian,l.kronologi_kejadian,"
                                +"l.tindakan_unit_kerja,l.tindakan_perbaikan,l.dibuat_oleh,l.disetujui_oleh,l.risiko_kredit"
                                +" From led_led l join led_ledlaporan ll on l.id = ll.led_laporan_id join led_laporan lap on lap.id = ll.laporan_id  where lap.id = 17")
    print(ledLaporan)
    
    pwd = os.path.dirname(__file__)
    file = pwd + '\static\periode.xlsx'
    file= load_workbook(file)
    
    print(type(file))
    
    ledbBase = []
    for laporanBaru in ledLaporan:
        ledDict = {
                "nomor_kasus": laporanBaru.nomor_kasus,
                "jumlah_kasus": laporanBaru.jumlah_kasus,
                "nama_pembuat": laporanBaru.nama_pembuat,
                "status": laporanBaru.status,
                "kode_cabang": laporanBaru.kode_cabang.kode_cabang,
                "kode_produk": laporanBaru.kode_produk.kode_produk,
                "nama_penemu": laporanBaru.nama_penemu,
                "kode_ktg_kejadian": laporanBaru.kode_ktg_kejadian.kode_kategori3,
                "kode_penyebab": laporanBaru.kode_penyebab.kode_penyebab3,
                "business_line": laporanBaru.business_line.kode_bisnis,
                "coa_biaya": laporanBaru.coa_biaya.kode_coa,
                "mata_uang": laporanBaru.mata_uang,
                "nilai_tukar": laporanBaru.nilai_tukar,
                "kerugian_potensial": laporanBaru.kerugian_potensial,
                "rra": laporanBaru.rra,
                "recovery": laporanBaru.recovery,
                "kerugian_aktual": laporanBaru.kerugian_aktual,
                "tanggal_review": laporanBaru.tanggal_review,
                "reviewer": laporanBaru.reviewer,
                "hasil_review": laporanBaru.hasil_review,
                "keterangan_review":laporanBaru.keterangan_review,
                "summary_kejadian": laporanBaru.summary_kejadian,
                "kronologi_kejadian": laporanBaru.kronologi_kejadian,
                "tindakan_unit_kerja": laporanBaru.tindakan_unit_kerja,
                "tindakan_perbaikan": laporanBaru.tindakan_perbaikan,
                "dibuat_oleh": laporanBaru.dibuat_oleh,
                "disetujui_oleh": laporanBaru.disetujui_oleh,
                "risiko_kredit": laporanBaru.risiko_kredit
        }
        ledbBase.append(ledDict)
        
        
        
    #import kategori kejadian
    #import bussines line
    #imprt organisasi
    #Import penyebab risiko
    #import led ormis
    #Import coa
    
    
    
    print(len(ledbBase))    
    # wb = Workbook()
    # ws = wb.active
    
    # cell = ws.cell(row=1,column=1)
    # cell.value = 'Id.'
    # cell.alignment = Alignment(horizontal='center', vertical='center')  
  
    # cell = ws.cell(row=1,column=2)
    # cell.value = 'Tanggal Kejadian'
    # cell.alignment = Alignment(horizontal='center', vertical='center')

    # cell = ws.cell(row=1,column=3)
    # cell.value = 'Tanggal Input'
    # cell.alignment = Alignment(horizontal='center', vertical='center')  

    # cell = ws.cell(row=1,column=4)
    # cell.value = 'Tanggal Terindentifikasi'
    # cell.alignment = Alignment(horizontal='center', vertical='center') 
    
    # cell = ws.cell(row=1,column=5)
    # cell.value = 'Tanggal Input'
    # cell.alignment = Alignment(horizontal='center', vertical='center')  
    
    # cell = ws.cell(row=1,column=6)
    # cell.value = 'Tanggal Selesai'
    # cell.alignment = Alignment(horizontal='center', vertical='center') 
    
    # cell = ws.cell(row=1,column=7)
    # cell.value = 'Tanggal Update'
    # cell.alignment = Alignment(horizontal='center', vertical='center')  
    
    # cell = ws.cell(row=1,column=8)
    # cell.value = 'Jumlah Kasus'
    # cell.alignment = Alignment(horizontal='center', vertical='center')    
    
    # cell = ws.cell(row=1,column=9)
    # cell.value = 'Tipe LED'
    # cell.alignment = Alignment(horizontal='center', vertical='center')   

    # cell = ws.cell(row=1,column=10)
    # cell.value = 'Status'
    # cell.alignment = Alignment(horizontal='center', vertical='center')   


    # cell = ws.cell(row=1,column=11)
    # cell.value = 'Kode Cabang'
    # cell.alignment = Alignment(horizontal='center', vertical='center')  
    
    # cell = ws.cell(row=1,column=12)
    # cell.value = 'Nama Cabang'
    # cell.alignment = Alignment(horizontal='center', vertical='center')    
    
    # cell = ws.cell(row=1,column=13)
    # cell.value = 'Kode Unit Kerja'
    # cell.alignment = Alignment(horizontal='center', vertical='center')  
    
    # cell = ws.cell(row=1,column=14)
    # cell.value = 'Nama Unit Kerja'
    # cell.alignment = Alignment(horizontal='center', vertical='center')    

    # cell = ws.cell(row=1,column=15)
    # cell.value = 'Nama Produk'
    # cell.alignment = Alignment(horizontal='center', vertical='center')   

    # cell = ws.cell(row=1,column=16)
    # cell.value = 'Ditemukan Oleh'
    # cell.alignment = Alignment(horizontal='center', vertical='center')  
    
    # cell = ws.cell(row=1,column=17)
    # cell.value = 'Kode Kategori Kejadian'
    # cell.alignment = Alignment(horizontal='center', vertical='center')    
    
    # cell = ws.cell(row=1,column=18)
    # cell.value = 'Kode Penyebab'
    # cell.alignment = Alignment(horizontal='center', vertical='center')  
    
    # cell = ws.cell(row=1,column=19)
    # cell.value = 'Business Line'
    # cell.alignment = Alignment(horizontal='center', vertical='center')    
    
    # cell = ws.cell(row=1,column=20)
    # cell.value = 'Coa Biaya'
    # cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    # cell = ws.cell(row=1,column=21)
    # cell.value = 'Mata Uang'
    # cell.alignment = Alignment(horizontal='center', vertical='center')   

    # cell = ws.cell(row=1,column=22)
    # cell.value = 'Nilai Tukar'
    # cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    # cell = ws.cell(row=1,column=23)
    # cell.value = 'Kerugian potensial'
    # cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    # cell = ws.cell(row=1,column=24)
    # cell.value = 'RRA'
    # cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    # cell = ws.cell(row=1,column=25)
    # cell.value = 'Recovery'
    # cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    # cell = ws.cell(row=1,column=26)
    # cell.value = 'Kerugian Aktual'
    # cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    # cell = ws.cell(row=1,column=27)
    # cell.value = 'Tanggal Review'
    # cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    # cell = ws.cell(row=1,column=28)
    # cell.value = 'Reviewer'
    # cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    # cell = ws.cell(row=1,column=29)
    # cell.value = 'Hasil Review'
    # cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    # cell = ws.cell(row=1,column=30)
    # cell.value = 'Keterangan Review'
    # cell.alignment = Alignment(horizontal='center', vertical='center') 
    
    # cell = ws.cell(row=1,column=31)
    # cell.value = 'Summary Kejadian'
    # cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    # cell = ws.cell(row=1,column=32)
    # cell.value = 'Kronologo Kejadian'
    # cell.alignment = Alignment(horizontal='center', vertical='center')     
    
    # cell = ws.cell(row=1,column=33)
    # cell.value = 'Tindakan Unit Kerja'
    # cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    # #looping from row 1
    # idrow= 2
    # idcol = 1

    # for data in ledLaporan:       
    #     #deskripsi identifikasi
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = data.id
    #     idcol+=1
   
    #     #likelihood
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = data.nomor_kasus
    #     idcol+=1
    #     #dampak
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = data.tanggal_kejadian
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = data.tanggal_teridentifikasi
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = data.tanggal_input
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = data.tanggal_update
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = data.jumlah_kasus
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = ""
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = data.status
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = str(data.kode_cabang)
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = "data.nama_cabang"
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = ""
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = ""
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = ""
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = ""
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = data.nama_penemu
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = str(data.kode_ktg_kejadian)
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = str(data.kode_penyebab)
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value =str(data.bussines_line)
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = str(data.coa_biaya)
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = data.mata_uang
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = data.nilai_tukar
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = data.kerugian_potensial
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = data.rra
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = data.recovery
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = data.kerugian_aktual
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = data.tanggal_review
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = data.keterangan_review
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = data.summary_kejadian
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = data.kronologi_kejadian
    #     idcol+=1
        
    #     cell = ws.cell(row=idrow,column=idcol)
    #     cell.value = data.tindakan_unit_kerja
    #     idcol+=1
        
    #     idrow += 1
    #     idcol =1
    
    
    # print(idrow) 
           
      
    # wb.save(response)

    # wb.close()
    
    # df = pd.DataFrame(list(LedLaporan.values()))
    df1 = pd.DataFrame([item.__dict__ for item in ledLaporan])
    print(df1)
    print("xx")
    dataframe1 = pd.read_excel(pwd + '\static\periode.xlsx')
    print(dataframe1)
    mydict = {
        "df": df1.to_html()
    }
    return render(request, 'led/dfTest.html', context=mydict)
    
def deleteLED(request,ledId):
    print("something")
    getLed = Led.objects.get(id = ledId)
    getLed.delete()
    
#masukkan ke log statechange dihapus
def deleteStateChange(sc):
    print("something")
    getLed = StateChange.objects.get(id = sc)
    print(getLed)
    getLed.delete()
       
def deleteStateChange2(request,ledId,scId):
    print("something")
    getLed = StateChange.objects.get(id = scId)
    print(getLed)
    getLed.delete() 
    return redirect('reviewLedId',ledId = ledId)  

def searchLedAjx(request):
    kd_cabang = Cabang.objects.all()
    kd_unit = Unit.objects.all()
    kd_produk = Produk.objects.all()
    kt_kejadian = KategoriKejadian.objects.all()
    ls_penyebab = Penyebab.objects.all()
    ls_bl   = BusinessLine.objects.all()
    print(ls_penyebab)
    context={
        'kd_cabang':kd_cabang,
        'kd_unit':kd_unit,
        'kd_produk':kd_produk,
        'kd_kejadian':kt_kejadian,
        'kd_penyebab':ls_penyebab,
        'kd_bl':ls_bl
    }
    if request.POST:
        print(dict(request.POST.items()))
        kotax = request.POST.getlist('kotax[]',False)
        kotax2 = request.POST.getlist('kode_cabang','')
        periodeAwal = request.POST.get('start_prd','')
        periodeAkhir = request.POST.get('end_prd','')
        tipeLed = request.POST.get('tipe_led','')
        status = request.POST.getlist('status','')
        kodeCabang = request.POST.getlist('kode_cabang','')
        produk = request.POST.get('produk','')
        kategoriKejadian = request.POST.getlist('kategori_kejadian','')
        kodePenyebab = request.POST.getlist('kode_penyebab','')
        businessLine = request.POST.getlist('business_line','')
        kerugianPotensial = request.POST.get('kerugian_potensial','')
        kerugianPotensialMin = request.POST.get('kerugian_potensial_min','')
        recovery = request.POST.get('recovery','')
        kerugianAktual = request.POST.get('kerugian_aktual','')
        reviewer =  request.POST.get('reviewer','')
        hasilReview = request.POST.get('hasil_review','')
        summaryKejadian = request.POST.get('summary_kejadian','')
        risikoKredit= request.POST.get('risiko_kredit','')

        
        #change date format
        periodeAwal = datetime.strptime(periodeAwal,'%Y-%m-%d')
        periodeAkhir = datetime.strptime(periodeAkhir,'%Y-%m-%d')  
        getNewLed = Led.objects.filter(tanggal_input__range =(periodeAwal,periodeAkhir)).select_related('coa_biaya','kode_produk','kode_ktg_kejadian','kode_penyebab','business_line','kode_cabang')
        addDict = {'data':getNewLed}
        context.update(addDict)
        #check value
        print("ss")
        print(kodeCabang)
        if( len(tipeLed) !=0 ):
            print("tipeLed")
            getNewLed = Led.objects.filter(tanggal_input__range =(periodeAwal,periodeAkhir)).filter(Q(tipe_led__iexact=tipeLed))
        print(len(status))
        if(status != False and len(status)!= 0):
            print("status")
            print("masuk")
            list_of_Q = [Q(**{'status__icontains': val}) for val in status]
            getNewLed = getNewLed.filter(reduce(operator.or_, list_of_Q))
            # for x in status:
            #     print(x)
            #     searchField = "status_icontains"
            #     getNewLed = getNewLed.filter(reduce)
        if(len(kodeCabang) !=0):
            print("kodeCABANG")
            getNewLed = getNewLed.filter(kode_cabang__in=kodeCabang)       
        if(len(produk) !=0):
            getNewLed = getNewLed.filter(kode_produk__in=produk)   
        if(len(kategoriKejadian) !=0):
            getNewLed = getNewLed.filter(kode_ktg_kejadian__in=kategoriKejadian)      
        if(len(kodePenyebab) !=0):
            getNewLed = getNewLed.filter(kode_penyebab__in=kodePenyebab)     
        if(len(businessLine) !=0):
            getNewLed = getNewLed.filter(business_line__in = businessLine)    
        if(kerugianPotensial !=''):
            getNewLed = getNewLed.filter(kerugian_potensial__gte = kerugianPotensial)
        if(kerugianPotensial !=''):
            getNewLed = getNewLed.filter(kerugian_potensial__lte = kerugianPotensialMin)
        if(recovery != ''):
            getNewLed = getNewLed.filter(recovery = recovery)
        if(kerugianAktual != ''):
            getNewLed= getNewLed.filter(aktual = kerugianAktual)
        if(reviewer != '' or len(reviewer)!= 0):
            getNewLed = getNewLed.filter(reviewer__icontains = reviewer)
        if(hasilReview != ''or len(hasilReview) != 0 ):
            getNewLed = getNewLed.filter(hasil_review__icontains = hasilReview)
        if(summaryKejadian != '' or len(summaryKejadian) != 0):
            getNewLed = getNewLed.filter(summary_kejadian__icontains = summaryKejadian)
        if(risikoKredit != ''or len(risikoKredit) != 0):
            getNewLed = getNewLed.filter(risiko_kredit__icontains = risikoKredit)
        paginator = Paginator(getNewLed,20)
        pages_number = request.GET.get('page')
        page_obj = Paginator.get_page(paginator,pages_number)
        
        addDict = {'page_obj':page_obj,
                   'lenData':len(getNewLed)}
        
        
        
        context.update(addDict)
        # print(context)
        return render(request,'led/search/searchLedAjx.html',context)
        
    
    print(dict(request.POST.items()))   
    return render(request,'led/search/searchLedAjx.html',context)

def searchLed(request,lapId):
    searchWord = request.POST.get('searchWord',False)
    sortBy   = request.POST.get('sortBy',False)
    getLaporan = Laporan.objects.get(id = lapId)
    
    # query = Led.objects.prefetch_related('ledlaporan_set').filter()
     
    search_choices = {
        '1':  'all',
        '2': 'led_laporan__hasil_review__icontains',
        '3': 'led_laporan__nomor_kasus__icontains',
        '4': 'led_laporan__reviewer__icontains',
        '5': 'laporan_id',
        }
    search_field = search_choices.get(str(sortBy))
    print(search_field)
    query = LedLaporan.objects.select_related('led_laporan','laporan').filter(**{search_field: searchWord},laporan_id = lapId)
    print(query)
    print(dict(request.POST.items()))
    # query = LedLaporan.objects.select_related('led_laporan','laporan').filter(laporan_id = 17)
    # print(query)
    
    context ={
        "data" :query,
        "laporan":getLaporan,
        # "getDuplicateData":getDuplicateData,
    }

    return render(request,'led/ledRpt/ledReport.html',context)
    
    # if(searchWord != False):
    #     print("#1")
    #     query = Led.objects.prefetch_related('ledlaporan_Set')
    #     print(query)
    #     #serializers
        
    # elif(searchWord !=  False and namaKolom != None):
    #     print("#2")     
     
    #if searchWord exist default cokumn is
    
    #if searchword and kolom
    
def deleteLedBySystems(request):
    # getDeleted = Led.objects.filter(Q(tanggal_kejadian__isnull) & Q(tanggal_teridentifikasi__isnull))
    # led pernah di review 
    #(dihapus == false) karena semu data baru false 
    getDeleted = Led.objects.filter(Q(tanggal_kejadian__isnull = True)&Q(tanggal_teridentifikasi__isnull = True)& Q(dihapus=False)).values_list('id')
    #led belum pernah di review
    unreviewedLed = Led.objects.filter(id__in = getDeleted)
    updateColumn = unreviewedLed.update(dihapus = True)
    
    #update column "dihapus" menjadi true pada data dihapus agar tidak digunakan 2 kali
    updateColumn = getDeleted.update(dihapus = True)
    
    #update column
    
    # led p
    data = serializers.serialize('json', unreviewedLed)
    
    return JsonResponse(data, safe=False)


def deleteLed(request):
    print("x")
    
def createLedHistory(request,lapId):
#buat
    # dt = datetime.now()
    # dt= dt.strftime("%m-%d-%Y %H:%M:%S")
    # print(dt)
    # print("x")
#time hisstory created
#nama laporan,tipe laporan,
    #laporan
    my_custom_sql(self =lapId )
            
    #copy to led master
    #my_custom_sql(self =lapId )
def my_custom_sql(self):
    print(self)
    lapId = self
    dt = timezone.localtime() 
    dt= dt.strftime("%Y-%m-%d %H:%M:%S")
    print(type(dt))
    print("x")
    
    histori = LedHistoris.objects.filter(Q(historis_date__isnull = False)).latest('historis_date')
    print(histori.historis_id)
    #buat id histori_id
    if(histori.historis_id == None):
        hsId = 1
    else:
        hsId = int(histori.historis_id) +1
    with connection.cursor() as cursor:
        # cursor.execute('INSERT INTO'+ '"ledMaster_ledhistoris"'+ '(laporan_id,historis_date,nomor_kasus,tanggal_kejadian,tanggal_teridentifikasi,tanggal_input,tanggal_selesai,tanggal_update,jumlah_kasus,nama_pembuat,status,kode_cabang,'+
		# 							'nama_cabang,kode_unit,nama_unit,tipe_led,kode_produk,nama_produk,nama_penemu,kode_ktg_kejadian1,nama_ktg_kejadian1,kode_ktg_kejadian2,nama_ktg_kejadian2,'+
		# 							 'kode_ktg_kejadian3,nama_ktg_kejadian3,kode_penyebab1,nama_penyebab1,kode_penyebab2,nama_penyebab2,kode_penyebab3,nama_penyebab3,bussines_line,'+
		# 							 'coa_biaya,mata_uang,nilai_tukar,kerugian_potensial,rra,kerugian_aktual,tanggal_review,reviewer,hasil_review,keterangan_review,summary_kejadian'+
		# 							 ',kronologi_kejadian,tindakan_unit_kerja,tindakan_perbaikan,dibuat_oleh,ditemukan_oleh,disetujui_oleh,dihapus,risiko_kredit,created_at,updated_at)'+
        # ' SELECT lp.laporan_id,'+"'"+dt+"'"+',l.nomor_kasus,l.tanggal_kejadian,l.tanggal_teridentifikasi,l.tanggal_input,l.tanggal_selesai,l.tanggal_update,l.jumlah_kasus,l.nama_pembuat,l.status,l.kode_cabang_id,'+
		# 							 'cb.nama_cabang,l.kode_unit_id,ut.nama_unit,l.tipe_led,l.kode_produk_id,prd.nama_produk,l.nama_penemu,kode_ktg_kejadian_id,nama_kategori1,kode_ktg_kejadian_id,'+
		# 							 'nama_kategori2,kode_ktg_kejadian_id,nama_kategori3,kode_penyebab_id,nama_penyebab1,kode_penyebab_id,nama_penyebab2,kode_penyebab_id,nama_penyebab3,bussines_line_id,'+
		# 							 'coa_biaya_id,mata_uang,nilai_tukar,kerugian_potensial,rra,kerugian_aktual,tanggal_review,reviewer,hasil_review,keterangan_review,summary_kejadian'+
		# 							 ',kronologi_kejadian,tindakan_unit_kerja,tindakan_perbaikan,dibuat_oleh,ditemukan_oleh,disetujui_oleh,dihapus,risiko_kredit,now(),l.updated_at '+
        # 'from led_led l join led_ledlaporan lp on lp.led_laporan_id = l.id '+
        # 'left join led_penyebab pb on pb.id = l.kode_penyebab_id '+
        # 'left join led_kategorikejadian kk on kk.id = l.kode_ktg_kejadian_id '+
        # 'left join led_cabang cb on cb.id = l.kode_cabang_id '+
        # 'left join led_unit ut on ut.id = l.kode_unit_id '+
        # 'left join led_coa coa on coa.id = l.coa_biaya_id '+
        # 'left join led_produk prd on prd.id = l.kode_produk_id ' +
        # 'where lp.laporan_id= '+ str(lapId))
        
        cursor.execute(
            "do $$ "+ 
                "declare "+
                "histori_date    timestamp := now(); "+
                "begin "+ 
                "INSERT INTO"+ '"'+"ledMaster_ledhistoris"+'"'+"(laporan_id,historis_id,historis_date,nomor_kasus,tanggal_kejadian,tanggal_teridentifikasi "+
                                                ",tanggal_input,tanggal_selesai,tanggal_update,jumlah_kasus,nama_pembuat,status "+
                                                ",kode_cabang,nama_cabang,kode_unit,nama_unit,tipe_led,kode_produk,nama_produk "+
                                                ",nama_penemu,kode_ktg_kejadian1,nama_ktg_kejadian1,kode_ktg_kejadian2,nama_ktg_kejadian2,kode_ktg_kejadian3 "+
                                                ",nama_ktg_kejadian3,kode_penyebab1,nama_penyebab1,kode_penyebab2,nama_penyebab2,kode_penyebab3,nama_penyebab3 "+
                                                ",bussines_line,coa_biaya,mata_uang,nilai_tukar,kerugian_potensial,rra,kerugian_aktual,tanggal_review,reviewer "+
                                                ",hasil_review,keterangan_review,summary_kejadian,kronologi_kejadian,tindakan_unit_kerja,tindakan_perbaikan "+
                                                ",dibuat_oleh,ditemukan_oleh,disetujui_oleh,dihapus,risiko_kredit,created_at,updated_at) "+
                                                "SELECT lp.laporan_id,"+str(hsId)+",histori_date,l.nomor_kasus,l.tanggal_kejadian,l.tanggal_teridentifikasi,l.tanggal_input "+
                                                ",l.tanggal_selesai,l.tanggal_update,l.jumlah_kasus,l.nama_pembuat,l.status,l.kode_cabang_id,cb.nama_cabang "+
                                                ",l.kode_unit_id,ut.nama_unit,l.tipe_led,l.kode_produk_id,prd.nama_produk,l.nama_penemu,kode_ktg_kejadian_id "+
                                                ",nama_kategori1,kode_ktg_kejadian_id,nama_kategori2,kode_ktg_kejadian_id,nama_kategori3,kode_penyebab_id "+
                                                ",nama_penyebab1,kode_penyebab_id,nama_penyebab2,kode_penyebab_id,nama_penyebab3,business_line_id "+
                                                ",coa_biaya_id,mata_uang,nilai_tukar,kerugian_potensial,rra,kerugian_aktual,tanggal_review "+
                                                ",reviewer,hasil_review,keterangan_review,summary_kejadian,kronologi_kejadian "+
                                                ",tindakan_unit_kerja,tindakan_perbaikan,dibuat_oleh,ditemukan_oleh,disetujui_oleh,dihapus "+
                                                ",risiko_kredit,now(),l.updated_at from led_led l join led_ledlaporan lp on lp.led_laporan_id = l.id left join led_penyebab pb on pb.id = l.kode_penyebab_id left join led_kategorikejadian kk on kk.id = l.kode_ktg_kejadian_id "+
                "left join led_cabang cb on cb.id = l.kode_cabang_id left join led_unit ut on ut.id = l.kode_unit_id "+
                "left join led_coa coa on coa.id = l.coa_biaya_id left join led_produk prd on prd.id = l.kode_produk_id where lp.laporan_id= "+str(lapId)+"; "+
                "end $$;"
            
        )
        row = cursor.fetchall()

    return row
    
    
def my_custom_sqlx(self):
    print(self)
    lapId = self
    dt = timezone.localtime() 
    dt= dt.strftime("%Y-%m-%d %H:%M:%S")
    print(type(dt))
    print("x")
    
    # histori = LedHistoris.objects.filter(Q(historis_date__isnull = False)).latest('historis_date')
    # print(histori.historis_id)
    # #buat id histori_id
    # if(histori.historis_id == 1):
    #     hsId = 1
    # else:
    #     hsId = int(histori.historis_id) +1
    with connection.cursor() as cursor:
       
        cursor.execute(
            "do $$ "+ 
                "declare "+
                "histori_date    timestamp := now(); "+
                "begin "+ 
                "INSERT INTO"+ '"'+"led_kasus"+'"'+"(nomor_kasus,tanggal_kejadian,tanggal_teridentifikasi "+
                                                ",tanggal_input,tanggal_selesai,tanggal_update,jumlah_kasus,nama_pembuat,status "+
                                                ",kode_cabang,nama_cabang,kode_unit,nama_unit,tipe_led,nama_produk "+
                                                ",nama_penemu,kode_ktg_kejadian,"+
                                                "kode_penyebab"+
                                                ",business_line,coa_biaya,mata_uang,nilai_tukar,kerugian_potensial,rra,kerugian_aktual,tanggal_review,reviewer "+
                                                ",hasil_review,keterangan_review,summary_kejadian,kronologi_kejadian,tindakan_unit_kerja,tindakan_perbaikan "+
                                                ",dibuat_oleh,ditemukan_oleh,disetujui_oleh,dihapus,chk_kknled,chk_summ,risiko_kredit,review_user,perlu_email,created_at,updated_at) "+
                                                "SELECT l.nomor_kasus,l.tanggal_kejadian,l.tanggal_teridentifikasi,l.tanggal_input "+
                                                ",l.tanggal_selesai,l.tanggal_update,l.jumlah_kasus,l.nama_pembuat,l.status,l.kode_cabang_id,cb.nama_cabang,"+
                                                "l.kode_unit_id,ut.nama_unit,l.tipe_led,prd.nama_produk,l.nama_penemu"+
                                                ",kode_ktg_kejadian_id,kode_penyebab_id,"+
                                                "business_line_id "+
                                                ",coa_biaya_id,mata_uang,nilai_tukar,kerugian_potensial,rra,kerugian_aktual,tanggal_review "+
                                                ",reviewer,hasil_review,keterangan_review,summary_kejadian,kronologi_kejadian "+
                                                ",tindakan_unit_kerja,tindakan_perbaikan,dibuat_oleh,ditemukan_oleh,disetujui_oleh,dihapus "+
                                                ",chk_kknled,chk_summ,risiko_kredit,review_user,perlu_email,now(),l.updated_at from led_led l join led_ledlaporan lp on lp.led_laporan_id = l.id left join led_penyebab pb on pb.id = l.kode_penyebab_id left join led_kategorikejadian kk on kk.id = l.kode_ktg_kejadian_id "+
                "left join led_cabang cb on cb.id = l.kode_cabang_id left join led_unit ut on ut.id = l.kode_unit_id "+
                "left join led_coa coa on coa.id = l.coa_biaya_id left join led_produk prd on prd.id = l.kode_produk_id where lp.laporan_id= "+str(lapId)+"; "+
                "end $$;"
            
        )
        row = cursor.fetchall()

    return row   
    
def saveLed(request,ledId):
    led = Led.objects.get(id = ledId)
    led.review_user = "Save"
    led.save()

    return True

def sumRevLed(request,getLaporan):
   
    rvBelum = LedLaporan.objects.filter(Q(laporan = getLaporan)).filter(Q(led_laporan__review_user__icontains = "Belum Review")).count()
    rvSave =    LedLaporan.objects.filter(Q(laporan = getLaporan)).filter(Q(led_laporan__review_user__icontains = "Save")).count()
    rvSubmit = LedLaporan.objects.filter(Q(laporan = getLaporan)).filter(Q(led_laporan__review_user__icontains = "Submit")).count()
    
    # print(type(getLaporan))
    getLaporan.ks_belumrv = rvBelum
    getLaporan.ks_save = rvSave
    getLaporan.ks_submit = rvSubmit
    
    getLaporan.save()
    
    # lsRv=[]
    # lsRv.append(rvBelum)
    # lsRv.append(rvSave)
    # lsRv.append(rvSubmit)
    
    # print(type(rvBelum))
    # print(rvSave)
    # print(rvSubmit)
    #belum review
    print("")
    #save
    #submit
    return True

def chkKategorixTipe(request,laporan):
    print("s")
    #get led per lapporan
    #count led
    data = LedLaporan.objects.select_related('laporan','led_laporan','led_laporan__coa_biaya','led_laporan__kode_produk','led_laporan__kode_ktg_kejadian','led_laporan__kode_penyebab').filter(laporan =laporan)

    #consider submit/save
    #consider tanggal update dari update ormis
    #apakah yang submit tidak perlu di check agar meringankan
    
    #sediakan tombol untuk menjalankan function
    for i in data:
        led = Led.objects.get(id = i.led_laporan.id)
        if(i.led_laporan.kode_ktg_kejadian.kode_kategori1 == '01' or i.led_laporan.kode_ktg_kejadian.kode_kategori1 =='02') : 
            if(i.led_laporan.tipe_led == 'NF' or i.led_laporan.tipe_led =='NFNL'):
    
                led.chk_kknled = True
                led.perlu_email = 1
                print("if")
                led.save()
           
        elif((i.led_laporan.kode_ktg_kejadian.kode_kategori1 != '01' and i.led_laporan.kode_ktg_kejadian.kode_kategori1 !='02')   ):
            if(i.led_laporan.tipe_led == 'FNL' or i.led_laporan.tipe_led =='FL'):
               
                led.chk_kknled = True
                led.perlu_email = 1
                print("elif")
                led.save()
        elif(i.chk_kknled == True):
            led.chk_kknled = False
            print("else")
            led.save()
    #loop
        #if kategorikejadian lvl 1 == 01/02 dan tipe led != FNL\FL
            #perlu diemail
        #if kategori kejadian != 01/02 dan tipe LED FL\FNL
            #perlu diemail         
#import

def checkSmr(request,laporan):
    for i in laporan:
        ledx = Led.objects.get(id = i.led_laporan.id)
        print(ledx.id)
        if(i.led_laporan.summary_kejadian == None):
            ledx.perlu_email = 1
            print("if chceksmr1")
            ledx.save()
       
            
#ada status lain selain blank?
def checkStatus(request,laporan):
    for i in laporan:
        led = Led.objects.get(id = i.led_laporan.id)
        if(len(i.led_laporan.status) == 0):
            led.perlu_email = 1
            print("check status")
            led.save()
 
def importAll(request):
    # xlsToXlsx(request)
    print("importAll")
    importLed(request)
    #importKasus(request)
    #importKtgKejadian(request)
    # importJrl(request)
    # importBL(request)
    # importCoa(request)
    # importPnyRisiko(request) 
    # importOrg(request)
    # importPrd(request)
    # importSummKejadian(request)

def importLed(request):
    pwd = os.path.dirname(__file__)
    file = pwd + '\static\TRX_REVIEW_LED_EDIT.xlsx'
    wb = load_workbook(file)
    sheets= wb.sheetnames
    ws=wb[sheets[2]]
    # ws = wb2.active
    dataImpor=[]
    print(ws.max_row)
    for i in range(2,ws.max_row+1):
        if ws.cell(row=i, column=1).value is None:
            break
        print("\n")
        dataImpor=[]
        print("Row ", i, " data :")
        for j in range(1,ws.max_column+1):
            cell_obj = ws.cell(row=i, column=j)
            
            dataImpor.append(cell_obj.value)
        #check aktif
        # if aktif =1 else 0
               
        led=Led()
        led.nomor_kasus= dataImpor[0]
        led.tanggal_kejadian=datetime.strptime(dataImpor[1],"%d/%m/%Y")
        led.tanggal_teridentifikasi=datetime.strptime(dataImpor[2],"%d/%m/%Y")
        led.tanggal_input=datetime.strptime(dataImpor[3],"%d/%m/%Y")
        print(dataImpor[4])
        if dataImpor[4]!=None:
            led.tanggal_selesai=datetime.strptime(dataImpor[4],"%d/%m/%Y")
            
        led.tanggal_update=dataImpor[5]
        led.jumlah_kasus=dataImpor[6]
        led.tipe_led=dataImpor[7]
        led.status= dataImpor[8]
        
        kd = Cabang.objects.filter(kode_cabang__iexact = str(dataImpor[9])).first()
        if kd:
            led.kode_cabang= kd
        else:
            led.kode_cabang= None
        led.ditemukan_oleh = dataImpor[14]
        
        #kk.nama_cabang =dataImpor[10]
        
       
        led.kode_unit= dataImpor[11]
       
        led.ditemukan_oleh = dataImpor[14]
        
        #kk.nama_unit_kerja = dataImpor[12]
        
        kd = Produk.objects.filter(nama_produk__iexact = str(dataImpor[13])).first()
        if kd:
            led.kode_produk= kd
        else:
            led.kode_produk= None
        led.ditemukan_oleh = dataImpor[14]
        
        try:
            kd = KategoriKejadian.objects.filter(kode_kategori3__iexact = str(dataImpor[15])).first()
        except kd.DoesNotExist:
            kd= None
        
        led.kode_ktg_kejadian=kd
        kd = Penyebab.objects.filter(kode_penyebab3__iexact = str(dataImpor[21])).first()
        if kd:
            led.kode_penyebab= kd
        else:
            led.kode_penyebab= None
        
        kd = BusinessLine.objects.filter(nama__iexact = str(dataImpor[27])).first()
        if kd:
            led.business_line= kd
        else:
            led.business_line= None
        
        kd = Coa.objects.filter(kode_coa__iexact = str(dataImpor[28])).first()
        if kd:
            led.coa_biaya= kd
        else:
            led.coa_biaya= None
            
        led.mata_uang = dataImpor[29]
        led.nilai_tukar= dataImpor[30]
        led.kerugian_potensial=dataImpor[31]
        led.rra=dataImpor[32]
        led.recovery= dataImpor[33]
        led.kerugian_aktual=dataImpor[34]
        if dataImpor[35]!=None:
            if isinstance(dataImpor[35],datetime):
                led.tanggal_review=dataImpor[35]
            else:
                led.tanggal_review= datetime.strptime(dataImpor[35],"%d/%m/%Y")
        else:
            led.tanggal_review= None
        # 
        # 
        
        led.reviewer = dataImpor[36]
        led.hasil_review = dataImpor[37]
        led.keterangan_review= dataImpor[38]
        led.summary_kejadian= dataImpor[39]
        led.kronologi_kejadian= dataImpor[40]
        
        led.tindakan_unit_kerja= dataImpor[41]
        led.tindakan_perbaikan= dataImpor[42]
        led.dibuat_oleh= dataImpor[43]
        led.disetujui_oleh= dataImpor[44]
        led.risiko_kredit= dataImpor[45]
        led.save()
    
def importKasus(request):
    pwd = os.path.dirname(__file__)
    file = pwd + '\static\TRX_REVIEW_LED_EDIT.xlsx'
    wb = load_workbook(file)
    sheets= wb.sheetnames
    ws=wb[sheets[1]]
    # ws = wb2.active
    dataImpor=[]
    print(ws.max_row)
    for i in range(2,ws.max_row+1):
        if ws.cell(row=i, column=1).value is None:
            break
        print("\n")
        dataImpor=[]
        print("Row ", i, " data :")
        for j in range(1,ws.max_column+1):
            cell_obj = ws.cell(row=i, column=j)
            
            dataImpor.append(cell_obj.value)
        #check aktif
        # if aktif =1 else 0
               
        led=Kasus()
        led.nomor_kasus= dataImpor[0]
        led.tanggal_kejadian=datetime.strptime(dataImpor[1],"%d/%m/%Y")
        led.tanggal_teridentifikasi=datetime.strptime(dataImpor[2],"%d/%m/%Y")
        led.tanggal_input=datetime.strptime(dataImpor[3],"%d/%m/%Y")
        print(dataImpor[4])
        if dataImpor[4]!=None:
            led.tanggal_selesai=datetime.strptime(dataImpor[4],"%d/%m/%Y")
        else:
            led.tanggal_selesai=None
        if dataImpor[5]!=None:
            led.tanggal_update=dataImpor[5]
        else:
            led.tanggal_update= None
        led.jumlah_kasus=dataImpor[6]
        led.tipe_led=dataImpor[7]
        led.status= dataImpor[8]
        led.kode_cabang= dataImpor[9]
        led.kode_unit= dataImpor[11]
        led.ditemukan_oleh = dataImpor[14]
        led.kode_ktg_kejadian=dataImpor[15]
        led.kode_penyebab=dataImpor[21]
        led.business_line=dataImpor[27]
        led.coa_biaya =dataImpor[28]
        led.mata_uang = dataImpor[29]
        led.nilai_tukar= dataImpor[30]
        led.kerugian_potensial=dataImpor[31]
        led.rra=dataImpor[32]
        led.recovery= dataImpor[33]
        led.kerugian_aktual=dataImpor[34]
        if dataImpor[35]!=None:
            led.tanggal_review=led.tanggal_review= datetime.strptime(dataImpor[35],"%d/%m/%Y")
        else:
            led.tanggal_review= None
        led.reviewer = dataImpor[36]
        led.hasil_review = dataImpor[37]
        led.keterangan_review= dataImpor[38]
        led.summary_kejadian= dataImpor[39]
        led.kronologi_kejadian= dataImpor[40]
        led.tindakan_unit_kerja= dataImpor[41]
        led.tindakan_perbaikan= dataImpor[42]
        led.dibuat_oleh= dataImpor[43]
        led.disetujui_oleh= dataImpor[44]
        led.risiko_kredit= dataImpor[45]
        
        led.save()
def importKtgKejadian(request):
    print("ktg")
  
       #ambil dari yang tanggal updatenya paling baru
    
    # xlsToXlsx(request)
    #get document from local
    #read
    #loop  untul null row
        #save
        #convert update date (date from ormis)
            #save update date
    KategoriKejadian.objects.all().delete()
    pwd = os.path.dirname(__file__)
    file = pwd + '\static\REF_KATEGORI_KEJADIAN.xlsx'
    wb = load_workbook(file)
    ws = wb.active
    dataImpor=[]
    print(ws.max_row)
    for i in range(2,ws.max_row+1):
        if ws.cell(row=i, column=1).value is None:
            break
        print("\n")
        dataImpor=[]
        print("Row ", i, " data :")
        for j in range(1,ws.max_column+1):
            cell_obj = ws.cell(row=i, column=j)
            dataImpor.append(cell_obj.value)
        #check aktif
        # if aktif =1 else 0
               
        kk=KategoriKejadian()
        kk.kode_kategori1= dataImpor[0]
        kk.nama_kategori1=dataImpor[1]
        kk.kode_kategori2=dataImpor[2]
        kk.nama_kategori2=dataImpor[3]
        kk.kode_kategori3=dataImpor[4]
        kk.nama_kategori3=dataImpor[5]
        kk.status = True
        kk.update_date = None
        kk.save()

def importSummKejadian(request):
    print("ksumm")
    
    
    #delete table sebelumnya dulu
    
    # xlsToXlsx(request)
    #get document from local
    #read
    #loop  untul null row
        #save
        #convert update date (date from ormis)
            #save update date
        
    pwd = os.path.dirname(__file__)
    file = pwd + '\static\SUMMARY_KEJADIAN.xlsx'
    wb = load_workbook(file)
    ws = wb.active
    dataImpor=[]
    
    for i in range(2,ws.max_row+1):
        print("\n")
        dataImpor=[]
        if ws.cell(row=i, column=1).value is None:
            break
        print("Row ", i, " data :")
        for j in range(1,ws.max_column+1):
            cell_obj = ws.cell(row=i, column=j)
            dataImpor.append(cell_obj.value)
        #check aktif
        # if aktif =1 else 0
               
        sk=SummaryKejadian()
        sk.kategori= dataImpor[0]
        sk.nama_lv1=dataImpor[1]
        sk.summ_ormis=dataImpor[2]
        sk.summ_ormc=dataImpor[3]
        sk.catatan=dataImpor[4]
        sk.save()


#BUSSINESS LINE
def importBL(request):
    print("importBl")
    
    # xlsToXlsx(request)
    #get document from local
    #read
    #loop  untul null row
        #save
        #convert update date (date from ormis)
            #save update date
        
    pwd = os.path.dirname(__file__)
    file = pwd + '\static\REF_BUSINESS_LINE.xlsx'
    wb = load_workbook(file)
    ws = wb.active
    dataImpor=[]
    
    for i in range(2,ws.max_row+1):
        print("\n")
        print("Row ", i, " data :")
        dataImpor=[]
        if ws.cell(row=i, column=1).value is None:
            break
        for j in range(1,ws.max_column+1):
            cell_obj = ws.cell(row=i, column=j)
            dataImpor.append(cell_obj.value)
        #check aktif
        # if aktif =1 else 0
               
        bl=BusinessLine()
        bl.kode_bisnis= dataImpor[0]
        bl.nama= dataImpor[1]
        bl.status=dataImpor[2]
        bl.update_date=dataImpor[3]
        bl.save()
        
#organisasi
def importOrg(request):
    print("importOrg")
    pwd = os.path.dirname(__file__)
    file = pwd + '\static\REF_ORGANISASI.xlsx'
    wb = load_workbook(file)
    ws = wb.active
    dataImpor=[]
    for i in range(2,ws.max_row+1):
        print("\n") 
        print("Row ", i, " data :")
        dataImpor=[]
        if ws.cell(row=i, column=1).value is None:
            break
        for j in range(1,ws.max_column+1):
            cell_obj = ws.cell(row=i, column=j)
            dataImpor.append(cell_obj.value)
        #check aktif
        # if aktif =1 else 0              
        org=Cabang()
        org.kode_cabang = dataImpor[0]
        org.tipe_cabang= dataImpor[1]
        org.nama_cabang=dataImpor[2]
        org.segmentasi  =dataImpor[3]
        org.status  =   dataImpor[4]
        org.update_date = None
        org.save()        
#penyebab risiko

def importPnyRisiko(request):
    print("importOrg")
    print("ktg")
    
    # xlsToXlsx(request)
    #get document from local
    #read
    #loop  untul null row
        #save
        #convert update date (date from ormis)
            #save update date
        
    pwd = os.path.dirname(__file__)
    file = pwd + '\static\REF_PENYEBAB_RISIKO.xlsx'
    wb = load_workbook(file)
    ws = wb.active
    #variable untuk status:
    aktif = True
    dataImpor=[]
    
    for i in range(2,ws.max_row+1):
        print("\n")
        dataImpor=[]
        if ws.cell(row=i, column=1).value is None:
            break
        print("Row ", i, " data :")
        for j in range(1,ws.max_column+1):
            cell_obj = ws.cell(row=i, column=j)
            dataImpor.append(cell_obj.value)
        #check aktif
 
            
        
        # if aktif =1 else 0
               
        pny=Penyebab()
        pny.kode_penyebab1= dataImpor[0]
        pny.nama_penyebab1=dataImpor[1]
        pny.kode_penyebab2=dataImpor[2]
        pny.nama_penyebab2=dataImpor[3]
        pny.kode_penyebab3=dataImpor[4]
        pny.nama_penyebab3=dataImpor[5]
        
        if(dataImpor[6] == "Aktif"):
            aktif= True 
        else:
            aktif=False
        pny.status = aktif
        # print(type(dataImpor[7]))
        # print(dataImpor[7])
        # print(datetime.now())
        pny.update_date = dataImpor[7]
        pny.save()   


#PRODUK    
 
def importPrd(request):
    print("Produk")
    # xlsToXlsx(request)
    #get document from local
    #read
    #loop  untul null row
        #save
        #convert update date (date from ormis)
            #save update date      
    pwd = os.path.dirname(__file__)
    file = pwd + '\static\REF_PRODUK.xlsx'
    wb = load_workbook(file)
    ws = wb.active
    dataImpor=[] 
    for i in range(2,ws.max_row+1):
        print("\n")
        dataImpor=[]
        if ws.cell(row=i, column=1).value is None:
            break
        print("Row ", i, " data :")
        for j in range(1,ws.max_column+1):
            cell_obj = ws.cell(row=i, column=j)
            dataImpor.append(cell_obj.value)
        #check aktif
        # if aktif =1 else 0 
        prd = Produk()
        prd.kode_produk = dataImpor[0]
        prd.tipe_produk = dataImpor[1]
        prd.nama_produk = dataImpor[2]
        prd.status = dataImpor[3]
        prd.update_date = None
        prd.save()   
        
        
#coa
 
def importCoa(request):
    print("COA")
    # xlsToXlsx(request)
    #get document from local
    #read
    #loop  untul null row
        #save
        #convert update date (date from ormis)
            #save update date      
    pwd = os.path.dirname(__file__)
    file = pwd + '\static\REF_KODE_COA.xlsx'
    wb = load_workbook(file)
    ws = wb.active
    dataImpor=[] 
    for i in range(2,ws.max_row+1):
        print("\n")
        dataImpor=[]
        if ws.cell(row=i, column=1).value is None:
            break
        print("Row ", i, " data :")
        for j in range(1,ws.max_column+1):
            cell_obj = ws.cell(row=i, column=j)
            dataImpor.append(cell_obj.value)
        #check aktif
        # if aktif =1 else 0 
        coa = Coa()
        coa.kode_coa = dataImpor[0]
        coa.nama = dataImpor[1]
        coa.status = dataImpor[2]
        coa.update_date = dataImpor[3]
        coa.save()   

# def importCoa(request):
#     print("COA")
#     # xlsToXlsx(request)
#     #get document from local
#     #read
#     #loop  untul null row
#         #save
#         #convert update date (date from ormis)
#             #save update date      
#     pwd = os.path.dirname(__file__)
#     file = pwd + '\static\REF_KODE_COA.xlsx'
#     wb = load_workbook(file)
#     ws = wb.active
#     dataImpor=[] 
#     for i in range(2,ws.max_row+1):
#         print("\n")
#         dataImpor=[]
#         if ws.cell(row=i, column=1).value is None:
#             break
#         print("Row ", i, " data :")
#         for j in range(1,ws.max_column+1):
#             cell_obj = ws.cell(row=i, column=j)
#             dataImpor.append(cell_obj.value)
#         #check aktif
#         # if aktif =1 else 0 
#         coa = Coa()
#         coa.kode_coa = dataImpor[0]
#         coa.nama = dataImpor[1]
#         coa.status = dataImpor[2]
#         coa.update_date = None
#         coa.save()   

def importJrl(request):
    print("Journal")
    # xlsToXlsx(request)
    #get document from local
    #read
    #loop  untul null row
        #save
        #convert update date (date from ormis)
            #save update date      
    pwd = os.path.dirname(__file__)
    file = pwd + '\static\TRX_JURNAL_REVIEW_LED.xlsx'
    wb = load_workbook(file)
    ws = wb.active
    dataImpor=[] 
    for i in range(2,ws.max_row+1):
        print("\n")
        dataImpor=[]
        if ws.cell(row=i, column=1).value is None:
            break
        print("Row ", i, " data :")
        for j in range(1,ws.max_column+1):
            cell_obj = ws.cell(row=i, column=j)
            dataImpor.append(cell_obj.value)
        #check aktif
        # if aktif =1 else 0 
        jrl = Journal()
        jrl.nomor_kasus= dataImpor[0]
        jrl.tanggal_jurnal = None
        
        jrl.debet_kredit = dataImpor[2]
        jrl.kode_mata_uang = dataImpor[3]
        jrl.jumlah_nominal_asal= float(dataImpor[4])
        jrl.jumlah_nominal_idr= float(dataImpor[5])
        jrl.keterangan= dataImpor[6]
        jrl.kode_coa = dataImpor [7]
        jrl.save()
     
#scheduler

def queryAllLed(request,rptId):
    data = Kasus.objects.raw("")
    
def listLedBySts(request,rptId):
    #check tipe
    # if(parTipe == "total"):
    #If else tipe
        #query sesuai tipe 
            #panggil function return query set
    context={
        "query":1,
        "reportId":1,
    }
    return render(request,'led/ledRpt/listLed.html',context)

def filterSearch(request):
    print("filterSearch")
    # led = Led.objects.all()
    led = Led.objects.select_related('kode_cabang','kode_produk','kode_penyebab','business_line','coa_biaya')
    led_filter = SearchFilter(request.GET, queryset =led)
    paginator = Paginator(led_filter.qs,20)
    pages_number = request.GET.get('page')
    page_obj = Paginator.get_page(paginator,pages_number)
    return render(request, 'led/listLed.html',{'page_obj':page_obj})
    
    