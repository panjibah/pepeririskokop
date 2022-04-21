
from django.shortcuts import render
from django.shortcuts import redirect
from .models import LedHistoris
from led.models import Laporan, Led,KategoriKejadian
from django.db.models import Q
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment 
from openpyxl import load_workbook
from django.http import HttpResponse
from django.utils import timezone

# Create your views here.

def listHistori(request):
    ledH = Laporan.objects.raw("SELECT ROW_NUMBER() OVER() AS numRow,historis_id, llap.id as id,llap.nama ,kl.nama as ktname,llap.created_at as tglLaporan,lhis.historis_date as tglHistori FROM "+'"ledMaster_ledhistoris"'+ " lhis join led_laporan llap on lhis.laporan_id = llap.id "
	+ "join led_kategorilaporan kl on kl.id = llap.kategori_id  group by llap.id,llap.nama,kl.nama,lhis.laporan_id,llap.created_at,lhis.historis_date,historis_id order by lhis.historis_date desc")
   
    context={
        
        "ledHistori":ledH,
    
    }
    
    return render(request,'ledMaster/histori/listLaporan.html',context)

def showHistori(request,lapId):
    # tglHistori = request.GET.get('tglHistori',False)
    hsId = request.GET.get('hsId',False)
    lapName = Laporan.objects.get(id = lapId)

    
    ledH = LedHistoris.objects.filter(Q(laporan_id = lapId)&Q(historis_id = hsId)).order_by('-tanggal_input')
    sizeData = ledH.count()
    print(ledH)
    context = {
        "ledHistori":ledH,
        "sizeData":sizeData,
        "hsId":hsId,
        "lapId":lapId,
        "lapName":lapName,
    }
    print(dict(request.POST.items()))
    return render(request,'ledMaster/histori/showHistori.html',context)

def exportHistori(request,lapId,hsId):
    laporan = Laporan.objects.get(id = lapId)

    #decide file name
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment;' + 'filename=' +  str(laporan.nama)+'.xlsx'
    
    x = "Risiko Operasional"
    ledLaporan= LedHistoris.objects.raw('Select id,l.nomor_kasus,l.tanggal_kejadian,l.tanggal_input,l.tanggal_teridentifikasi,l.tanggal_input,l.tanggal_selesai,l.jumlah_kasus,'
                                +'l.nama_pembuat,l.status,l.kode_cabang,l.kode_produk,l.nama_penemu,l.kode_ktg_kejadian3,l.kode_penyebab3,l.bussines_line,l.coa_biaya,l.mata_uang,l.nilai_tukar,l.kerugian_potensial,l.rra,'
                                +'l.recovery,l.kerugian_aktual,l.tanggal_review,l.reviewer,l.hasil_review,l.keterangan_review,l.summary_kejadian,l.kronologi_kejadian,'
                                +'l.tindakan_unit_kerja,l.tindakan_perbaikan,l.dibuat_oleh,l.disetujui_oleh,l.risiko_kredit'
                                +' From '+'"' +"ledMaster_ledhistoris" +'" l' +' where laporan_id ='+ str(lapId) +' and historis_id='+str(hsId))
    #takenRisiko = IdentifikasiRisiko.objects.raw("SELECT DISTINCT IR.id,LENGTH(idn_komentar) AS panjang_komentar,RA.inherent,idn_deskripsi,idn_likelihood,idn_dampak,idn_nilai_risiko,identifikasi_risiko_id,mtg_deskripsi,mtg_likelihood,mtg_dampak,mtg_nilai_risiko,mtg_keterangan_likelihood,mtg_keterangan_dampak FROM public.library_riskassessment RA JOIN library_identifikasirisiko IR ON RA.id = IR.idn_risk_assessment_id JOIN library_mitigasirisiko MR ON MR.identifikasi_risiko_id = IR.id JOIN library_tiperisiko tp ON IR.tipe_risiko_id = tp.id WHERE RA.id = %s AND tp.tipe_risiko_nama LIKE 'Risiko Hukum' ",[raID])
    # print(LedLaporan)
    wb = Workbook()
    ws = wb.active
    
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment;' + 'filename=' +  str(laporan.nama)+'.xlsx'
    #create heading
    cell = ws.cell(row=1,column=1)
    cell.value = 'Nomor Kasus'
    cell.alignment = Alignment(horizontal='center', vertical='center')  
  
    cell = ws.cell(row=1,column=2)
    cell.value = 'Tanggal Kejadian'
    cell.alignment = Alignment(horizontal='center', vertical='center')

    cell = ws.cell(row=1,column=3)
    cell.value = 'Tanggal Input'
    cell.alignment = Alignment(horizontal='center', vertical='center')  

    cell = ws.cell(row=1,column=4)
    cell.value = 'Tanggal Terindentifikasi'
    cell.alignment = Alignment(horizontal='center', vertical='center') 
    
    cell = ws.cell(row=1,column=5)
    cell.value = 'Tanggal Input'
    cell.alignment = Alignment(horizontal='center', vertical='center')  
    
    cell = ws.cell(row=1,column=6)
    cell.value = 'Tanggal Selesai'
    cell.alignment = Alignment(horizontal='center', vertical='center') 
    
    cell = ws.cell(row=1,column=7)
    cell.value = 'Tanggal Update'
    cell.alignment = Alignment(horizontal='center', vertical='center')  
    
    cell = ws.cell(row=1,column=8)
    cell.value = 'Jumlah Kasus'
    cell.alignment = Alignment(horizontal='center', vertical='center')    
    
    cell = ws.cell(row=1,column=9)
    cell.value = 'Tipe LED'
    cell.alignment = Alignment(horizontal='center', vertical='center')   

    cell = ws.cell(row=1,column=10)
    cell.value = 'Status'
    cell.alignment = Alignment(horizontal='center', vertical='center')   


    cell = ws.cell(row=1,column=11)
    cell.value = 'Kode Cabang'
    cell.alignment = Alignment(horizontal='center', vertical='center')  
    
    cell = ws.cell(row=1,column=12)
    cell.value = 'Nama Cabang'
    cell.alignment = Alignment(horizontal='center', vertical='center')    
    
    cell = ws.cell(row=1,column=13)
    cell.value = 'Kode Unit Kerja'
    cell.alignment = Alignment(horizontal='center', vertical='center')  
    
    cell = ws.cell(row=1,column=14)
    cell.value = 'Nama Unit Kerja'
    cell.alignment = Alignment(horizontal='center', vertical='center')    

    cell = ws.cell(row=1,column=15)
    cell.value = 'Nama Produk'
    cell.alignment = Alignment(horizontal='center', vertical='center')   

    cell = ws.cell(row=1,column=16)
    cell.value = 'Ditemukan Oleh'
    cell.alignment = Alignment(horizontal='center', vertical='center')  
    
    cell = ws.cell(row=1,column=17)
    cell.value = 'Kode Kategori Kejadian'
    cell.alignment = Alignment(horizontal='center', vertical='center')    
    
    cell = ws.cell(row=1,column=18)
    cell.value = 'Kode Penyebab'
    cell.alignment = Alignment(horizontal='center', vertical='center')  
    
    cell = ws.cell(row=1,column=19)
    cell.value = 'Business Line'
    cell.alignment = Alignment(horizontal='center', vertical='center')    
    
    cell = ws.cell(row=1,column=20)
    cell.value = 'Coa Biaya'
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    cell = ws.cell(row=1,column=21)
    cell.value = 'Mata Uang'
    cell.alignment = Alignment(horizontal='center', vertical='center')   

    cell = ws.cell(row=1,column=22)
    cell.value = 'Nilai Tukar'
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    cell = ws.cell(row=1,column=23)
    cell.value = 'Kerugian potensial'
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    cell = ws.cell(row=1,column=24)
    cell.value = 'RRA'
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    cell = ws.cell(row=1,column=25)
    cell.value = 'Recovery'
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    cell = ws.cell(row=1,column=26)
    cell.value = 'Kerugian Aktual'
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    cell = ws.cell(row=1,column=27)
    cell.value = 'Tanggal Review'
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    cell = ws.cell(row=1,column=28)
    cell.value = 'Reviewer'
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    cell = ws.cell(row=1,column=29)
    cell.value = 'Hasil Review'
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    cell = ws.cell(row=1,column=30)
    cell.value = 'Keterangan Review'
    cell.alignment = Alignment(horizontal='center', vertical='center') 
    
    cell = ws.cell(row=1,column=31)
    cell.value = 'Summary Kejadian'
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    
    cell = ws.cell(row=1,column=32)
    cell.value = 'Kronologo Kejadian'
    cell.alignment = Alignment(horizontal='center', vertical='center')     
    
    cell = ws.cell(row=1,column=33)
    cell.value = 'Tindakan Unit Kerja'
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    

    #looping from row 1
    idrow= 2
    idcol = 1

   
    for data in ledLaporan:       
        #deskripsi identifikasi
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.nomor_kasus
        idcol+=1
   
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.tanggal_kejadian
        idcol+=1
      
      
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value =data.tanggal_teridentifikasi
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value =data.tanggal_input
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value =data.tanggal_selesai
        idcol+=1
        
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = str(data.tanggal_update)
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.jumlah_kasus
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = ""
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.status
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = str(data.kode_cabang)
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.nama_cabang
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = ""
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = ""
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = ""
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = ""
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.nama_penemu
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = str(data.kode_ktg_kejadian3)
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = str(data.kode_penyebab3)
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value =str(data.bussines_line)
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = str(data.coa_biaya)
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.mata_uang
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.nilai_tukar
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.kerugian_potensial
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.rra
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.recovery
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.kerugian_aktual
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.tanggal_review
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.keterangan_review
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.summary_kejadian
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.kronologi_kejadian
        idcol+=1
        
        cell = ws.cell(row=idrow,column=idcol)
        cell.value = data.tindakan_unit_kerja
        idcol+=1
        
        idrow += 1
        idcol =1
    
    
    print(idrow) 
           
      
    wb.save(response)

    wb.close()
    context = {
    }
    return response